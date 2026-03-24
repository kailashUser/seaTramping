"""
SEA Tramping Voyage Simulation Dashboard V2
Network graph optimisation + exact costing model + voyage-level detail + What-If editor.
"""
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import sys
import os
import time
import math
import json as _json

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from modules.data_processor import (
    load_and_process_data, build_port_database, build_distance_matrix,
    build_leg_library, COMMODITY_23, COMMODITY_CATEGORIES, FREIGHT_RATE_PARAMS,
    PORT_CHARGES_BY_COUNTRY, HANDLING_COST
)
from modules.simulation_engine import (
    VesselConfig, SimConfig,
    run_full_simulation, analyse_results,
    build_voyage_graph, compute_port_centrality, find_communities,
    cascade_recalculate_legs, FastLegLibrary, _find_leg_idx,
    HAS_NX,
)

st.set_page_config(page_title="SEA Tramping Simulation V2", layout="wide", page_icon="🚢")

st.markdown("""
<style>
    .metric-card {background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.2rem; border-radius: 12px; color: white; text-align: center;}
    .metric-value {font-size: 1.8rem; font-weight: 700;}
    .metric-label {font-size: 0.85rem; opacity: 0.9;}
    .stTabs [data-baseweb="tab-list"] {gap: 8px;}
    .stTabs [data-baseweb="tab"] {padding: 10px 20px; font-weight: 600;}
    .profit-pos {color: #22c55e; font-weight: 700;}
    .profit-neg {color: #ef4444; font-weight: 700;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="
    background:linear-gradient(135deg,#1a3a5c 0%,#2d5986 60%,#667eea 100%);
    padding:20px 28px 16px;
    border-radius:12px;
    margin-bottom:24px;
">
    <div style="font-size:2rem;font-weight:700;color:white;letter-spacing:-0.5px;">
        🚢 SEA Tramping Voyage Optimiser
    </div>
    <div style="font-size:0.95rem;color:rgba(255,255,255,0.75);margin-top:4px;">
        Monte Carlo + Graph Optimisation &nbsp;|&nbsp;
        Dual Fuel Exact Costing &nbsp;|&nbsp;
        Voyage Dependency Cascade &nbsp;|&nbsp;
        70 Ports × 23 Commodities × 5 Years Data
    </div>
</div>
""", unsafe_allow_html=True)


@st.cache_data(ttl=3600)
def fetch_live_bunker_prices():
    """
    Fetch live bunker prices from Ship & Bunker (Singapore).
    Returns dict with LSFO and MGO prices, or None on failure.
    Falls back silently if network is unavailable.
    """
    try:
        import requests
        from bs4 import BeautifulSoup
        url = "https://shipandbunker.com/prices/apac/sea/sg-sin-singapore"
        headers = {'User-Agent': 'Mozilla/5.0 (compatible; research bot)'}
        r = requests.get(url, headers=headers, timeout=8)
        if r.status_code != 200:
            return None
        soup = BeautifulSoup(r.text, 'html.parser')
        prices = {}
        for row in soup.select('table tr'):
            cells = row.find_all('td')
            if len(cells) >= 2:
                fuel = cells[0].get_text(strip=True).upper()
                price_text = cells[1].get_text(strip=True).replace(',', '')
                try:
                    price = float(price_text)
                    if 'VLSFO' in fuel or 'LSFO' in fuel:
                        prices['lsfo'] = price
                    elif 'MGO' in fuel:
                        prices['mgo'] = price
                except ValueError:
                    pass
        return prices if prices else None
    except Exception:
        return None


def export_programme_to_excel(prog, vessel_config, programme_rank):
    """
    Export a voyage programme to a formatted Excel workbook matching the
    company tramping spreadsheet layout.
    Returns bytes object ready for st.download_button.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Programme #{programme_rank}"

    header_fill  = PatternFill("solid", fgColor="1a3a5c")
    subhead_fill = PatternFill("solid", fgColor="667eea")
    alt_fill     = PatternFill("solid", fgColor="f0f4ff")
    total_fill   = PatternFill("solid", fgColor="22c55e")
    white_font   = Font(color="FFFFFF", bold=True, size=10)
    bold_font    = Font(bold=True, size=10)
    normal_font  = Font(size=10)
    money_fmt    = '#,##0'
    rate_fmt     = '#,##0.00'
    thin_border  = Border(bottom=Side(style='thin', color='cccccc'))

    # Title block
    ws.merge_cells('A1:AC1')
    ws['A1'] = f"SEA Tramping Voyage Programme #{programme_rank} — Annual P&L"
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=13)
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:AC2')
    ws['A2'] = (
        f"Vessel DWT: {vessel_config.dwt:,} MT  |  "
        f"DWCC: {vessel_config.dwcc:,} MT  |  "
        f"Speed: {vessel_config.speed_laden_knots}/{vessel_config.speed_ballast_knots} kn  |  "
        f"Charter Hire: ${vessel_config.charter_hire_day:,}/day  |  "
        f"LSFO: ${vessel_config.lsfo_price_mt}/MT  |  "
        f"MGO: ${vessel_config.mgo_price_mt}/MT"
    )
    ws['A2'].font = Font(size=9, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    headers = [
        'Voy #', 'Ballast From', 'Load Port', 'Disch Port', 'Commodity',
        'Cargo MT', 'Rate $/MT', 'Gross Freight',
        'Brokerage', 'Net Income',
        'Charter Hire', 'LSFO MT', 'MGO MT', 'LSFO Cost', 'MGO Cost',
        'Total Bunker', 'Load Port Nav', 'Disch Port Nav', 'Port Costs',
        'Insurance', 'Other Costs', 'Total Expenses',
        'Profit / Loss', '$/Day', '$/MT',
        'Ballast NM', 'Laden NM', 'Ballast Days', 'Laden Days',
        'Port Days', 'Cong Days', 'Total Days', 'Cum Profit',
    ]
    col_widths = [
        6, 16, 16, 16, 18,
        10, 9, 14,
        11, 14,
        14, 9, 9, 13, 13,
        13, 14, 14, 13,
        11, 11, 14,
        14, 11, 9,
        11, 11, 12, 12,
        11, 11, 12, 14,
    ]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = white_font
        cell.fill = subhead_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = 'A4'

    legs = prog['legs']
    for ri, leg in enumerate(legs):
        row = 4 + ri
        fill = alt_fill if ri % 2 == 0 else None
        g   = leg.get('gross_freight', leg.get('revenue', 0))
        br  = leg.get('brokerage', g * 0.0375)
        ni  = leg.get('net_income', g - br)
        lc  = leg.get('lsfo_cost', 0)
        mc  = leg.get('mgo_cost', 0)
        tb  = lc + mc
        ch  = leg.get('charter_hire', leg.get('charter_hire_cost', 0))
        pc  = leg.get('port_costs',
              leg.get('load_port_nav', 0) + leg.get('disch_port_nav', 0))
        ins = leg.get('insurance', 0)
        oth = leg.get('other_costs', 1000)
        exp = leg.get('total_expenses', leg.get('total_cost', 0))
        pl  = leg.get('profit_loss', leg.get('profit', 0))
        ppd = leg.get('profit_per_day', 0)
        pmt = leg.get('profit_per_mt', 0)
        cum = leg.get('cum_profit', 0)

        values = [
            ri + 1,
            leg.get('ballast_from_port', '—'),
            leg.get('origin_port', ''),
            leg.get('dest_port', ''),
            leg.get('commodity', ''),
            leg.get('cargo_mt', 0),
            leg.get('freight_rate', 0),
            g, br, ni,
            ch,
            leg.get('lsfo_mt', 0),
            leg.get('mgo_mt', 0),
            lc, mc, tb,
            leg.get('load_port_nav', 0),
            leg.get('disch_port_nav', 0),
            pc, ins, oth, exp,
            pl, ppd, pmt,
            leg.get('ballast_nm', leg.get('ballast_distance_nm', 0)),
            leg.get('laden_nm', leg.get('distance_nm', 0)),
            leg.get('ballast_days', leg.get('sailing_days_ballast', 0)),
            leg.get('laden_days', leg.get('sailing_days_laden', 0)),
            leg.get('port_days', 0),
            leg.get('congestion_days', 0),
            leg.get('total_days', 0),
            cum,
        ]
        money_cols = {8, 9, 10, 11, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 33}
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='right' if ci > 5 else 'left')
            if fill:
                cell.fill = fill
            if ci in money_cols and isinstance(val, (int, float)):
                cell.number_format = money_fmt
            elif ci in {7, 24, 25} and isinstance(val, (int, float)):
                cell.number_format = rate_fmt
        pl_cell = ws.cell(row=row, column=23)
        if pl > 0:
            pl_cell.font = Font(color="166534", bold=True, size=10)
        elif pl < 0:
            pl_cell.font = Font(color="991b1b", bold=True, size=10)

    # Totals row
    tr = 4 + len(legs)
    ws.cell(row=tr, column=1, value='TOTAL').font = bold_font
    total_cols = {
        8:  sum(l.get('gross_freight', l.get('revenue', 0)) for l in legs),
        9:  sum(l.get('brokerage', 0) for l in legs),
        10: sum(l.get('net_income', 0) for l in legs),
        11: sum(l.get('charter_hire', l.get('charter_hire_cost', 0)) for l in legs),
        16: sum(l.get('lsfo_cost', 0) + l.get('mgo_cost', 0) for l in legs),
        19: sum(l.get('port_costs', 0) for l in legs),
        20: sum(l.get('insurance', 0) for l in legs),
        22: sum(l.get('total_expenses', l.get('total_cost', 0)) for l in legs),
        23: sum(l.get('profit_loss', l.get('profit', 0)) for l in legs),
        32: sum(l.get('total_days', 0) for l in legs),
    }
    for ci, val in total_cols.items():
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = total_fill
        cell.number_format = money_fmt if ci != 32 else '#,##0.0'

    # P&L Summary sheet
    ws2 = wb.create_sheet("Annual P&L Summary")
    total_gross  = sum(l.get('gross_freight', l.get('revenue', 0)) for l in legs)
    total_broker = sum(l.get('brokerage', 0) for l in legs)
    total_ni     = total_gross - total_broker
    total_hire   = sum(l.get('charter_hire', l.get('charter_hire_cost', 0)) for l in legs)
    total_lsfo   = sum(l.get('lsfo_cost', 0) for l in legs)
    total_mgo    = sum(l.get('mgo_cost', 0) for l in legs)
    total_port   = sum(l.get('port_costs', 0) for l in legs)
    total_ins    = sum(l.get('insurance', 0) for l in legs)
    total_other  = sum(l.get('other_costs', 1000) for l in legs)
    total_exp    = sum(l.get('total_expenses', l.get('total_cost', 0)) for l in legs)
    total_pl     = total_ni - total_exp
    total_days   = sum(l.get('total_days', 0) for l in legs)
    tce          = (total_ni - total_port - total_ins - total_other) / max(total_days, 1)

    pl_lines = [
        ("INCOME", None),
        ("Gross Freight Revenue", total_gross),
        ("  Less: Brokerage (3.75%)", -total_broker),
        ("Net Income", total_ni),
        ("", None),
        ("EXPENSES", None),
        ("Charter Hire", -total_hire),
        ("Bunker — LSFO", -total_lsfo),
        ("Bunker — MGO", -total_mgo),
        ("Total Bunker Cost", -(total_lsfo + total_mgo)),
        ("Port Navigation Costs", -total_port),
        ("Insurance", -total_ins),
        ("Other Costs", -total_other),
        ("Total Expenses", -total_exp),
        ("", None),
        ("NET PROFIT / LOSS", total_pl),
        ("", None),
        ("KPIs", None),
        ("TCE ($/day)", tce),
        ("Total Operating Days", total_days),
        ("Number of Voyages", len(legs)),
        ("Avg Profit per Voyage", total_pl / max(len(legs), 1)),
    ]
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 18
    for ri, (label, val) in enumerate(pl_lines, 1):
        lc = ws2.cell(row=ri, column=1, value=label)
        if val is None:
            lc.font = Font(bold=True, size=11)
            lc.fill = PatternFill("solid", fgColor="e8edf7")
        else:
            lc.font = Font(size=10)
            vc = ws2.cell(row=ri, column=2, value=val)
            vc.number_format = money_fmt
            vc.alignment = Alignment(horizontal='right')
            if label == "NET PROFIT / LOSS":
                lc.font = Font(bold=True, size=12)
                vc.font = Font(bold=True, size=12,
                               color="166534" if val >= 0 else "991b1b")
            elif label.startswith("TCE"):
                vc.number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@st.cache_data
def load_data(excel_path):
    intra = load_and_process_data(excel_path)
    ports = build_port_database(intra, n_ports=70)
    dist_matrix = build_distance_matrix(ports)
    legs = build_leg_library(ports, dist_matrix, intra)
    return intra, ports, dist_matrix, legs


# ─── SIMULATION STATE FLAGS ──────────────────────────────────────────────────
if 'sim_running' not in st.session_state:
    st.session_state['sim_running'] = False
if 'sim_done' not in st.session_state:
    st.session_state['sim_done'] = False
if 'sim_requested' not in st.session_state:
    st.session_state['sim_requested'] = False
_sim_locked = st.session_state.get('sim_running', False)

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
st.sidebar.markdown("## 🚢 Vessel Configuration")
dwt   = st.sidebar.number_input("Deadweight (DWT)", value=17556, min_value=5000, max_value=60000, step=500, disabled=_sim_locked)
dwcc  = st.sidebar.number_input("Cargo Capacity (DWCC, MT)", value=15000, min_value=3000, max_value=55000, step=500, disabled=_sim_locked)
speed_laden    = st.sidebar.number_input("Speed Laden (knots)",   value=11.0, min_value=8.0, max_value=16.0, step=0.5, disabled=_sim_locked)
speed_ballast  = st.sidebar.number_input("Speed Ballast (knots)", value=11.5, min_value=8.0, max_value=16.0, step=0.5, disabled=_sim_locked)

st.sidebar.markdown("## 💰 Cost Parameters")
charter_hire = st.sidebar.number_input(
    "Charter Hire (USD/day)", value=9000, min_value=3000, max_value=30000, step=500,
    disabled=_sim_locked
)

_live = fetch_live_bunker_prices()
_lsfo_default = int(_live['lsfo']) if _live and 'lsfo' in _live else 560
_mgo_default  = int(_live['mgo'])  if _live and 'mgo'  in _live else 780
_live_badge   = "🟢 Live price" if (_live and len(_live) >= 2) else "⚪ Using defaults"

st.sidebar.markdown(f"**Bunker Prices** {_live_badge}")
if _live and len(_live) >= 2:
    st.sidebar.caption(
        f"Live from Ship & Bunker Singapore: "
        f"VLSFO ${_live.get('lsfo', '—')} | MGO ${_live.get('mgo', '—')}"
    )
else:
    st.sidebar.caption(
        "Could not fetch live prices — using sidebar defaults. "
        "Edit manually above."
    )
lsfo_price = st.sidebar.number_input(
    "LSFO / VLSFO Price (USD/MT)", value=_lsfo_default, min_value=200, max_value=1200, step=10,
    disabled=_sim_locked
)
mgo_price = st.sidebar.number_input(
    "MGO Price (USD/MT)", value=_mgo_default, min_value=300, max_value=1500, step=10,
    disabled=_sim_locked
)
insurance_annual  = st.sidebar.number_input("Insurance (USD/year)",   value=16000, min_value=5000,  max_value=100000, step=1000, disabled=_sim_locked)
brokerage_pct_ui  = st.sidebar.number_input("Brokerage (%)",          value=3.75,  min_value=0.0,   max_value=10.0,   step=0.25, disabled=_sim_locked)
operating_days    = st.sidebar.number_input("Operating Days/Year",     value=330,   min_value=270,   max_value=365,    step=5,    disabled=_sim_locked)

st.sidebar.markdown("## 🏗️ Port Cost Settings")
fio_cargo    = st.sidebar.checkbox("FIO Cargo (Free In & Out — no stevedoring)", value=True,
                                    help="Uncheck to apply stevedoring charges per commodity rate",
                                    disabled=_sim_locked)
port_charge_vol = st.sidebar.slider("Port Charge Variation (±%)", 0, 30, 15, 5,
                                     help="Stochastic ±% variation applied to port navigation charges per iteration",
                                     disabled=_sim_locked) / 100.0
congestion_vol  = st.sidebar.slider("Congestion Variation (±%)", 0, 60, 40, 5,
                                     help="Stochastic ±% variation applied to waiting/congestion days per iteration",
                                     disabled=_sim_locked) / 100.0

with st.sidebar.expander("Port Charge Overrides (CSV)"):
    from data.port_charges import PORT_CHARGES, PORT_CHARGES_DEFAULT
    import io
    _pc_rows = []
    for pname, vals in PORT_CHARGES.items():
        _pc_rows.append({'Port': pname, 'Nav_USD': vals['nav'],
                         'Cong_Mean_Days': vals['cong_mean'], 'Cong_Std_Days': vals['cong_std']})
    _pc_df = pd.DataFrame(_pc_rows)
    _csv_bytes = _pc_df.to_csv(index=False).encode()
    st.download_button("Download Port Charges CSV", data=_csv_bytes,
                       file_name="port_charges_override.csv", mime="text/csv",
                       help="Edit and re-upload to override default port charges")
    _uploaded = st.file_uploader("Upload Modified CSV", type="csv", key="port_charge_upload")
    if _uploaded is not None:
        try:
            _override_df = pd.read_csv(_uploaded)
            _override_map = {row['Port']: {'nav': row['Nav_USD'],
                                            'cong_mean': row['Cong_Mean_Days'],
                                            'cong_std': row['Cong_Std_Days']}
                             for _, row in _override_df.iterrows()}
            PORT_CHARGES.update(_override_map)
            st.success(f"Loaded {len(_override_df)} port charge overrides.")
        except Exception as _e:
            st.error(f"Invalid CSV: {_e}")

st.sidebar.markdown("---")
st.sidebar.markdown("## 🚀 Simulation")

algo_choice = st.sidebar.selectbox(
    "Algorithm",
    ["Hybrid (Greedy + Monte Carlo)", "Monte Carlo Only", "Greedy + Local Search"],
    index=0,
    key='algo_choice_sidebar',
    disabled=_sim_locked,
)
n_iterations = st.sidebar.selectbox(
    "Iterations",
    [1000, 5000, 10000, 25000, 50000],
    index=2,
    key='n_iterations_sidebar',
    disabled=_sim_locked,
)
freight_vol = st.sidebar.slider(
    "Freight Rate Volatility", 0.05, 0.30, 0.15, 0.05,
    key='freight_vol_sidebar',
    disabled=_sim_locked,
)
bunker_vol = st.sidebar.slider(
    "Bunker Price Volatility", 0.05, 0.25, 0.10, 0.05,
    key='bunker_vol_sidebar',
    disabled=_sim_locked,
)

st.sidebar.markdown("<br>", unsafe_allow_html=True)
if _sim_locked:
    st.sidebar.markdown(
        "<div style='background:linear-gradient(135deg,#f59e0b22,#f59e0b44);"
        "border-left:4px solid #f59e0b;padding:10px 14px;border-radius:8px;"
        "font-size:0.85rem;color:#92400e;font-weight:600'>"
        "⏳ Simulation running...<br>"
        "<span style='font-weight:400;font-size:0.78rem'>"
        "Controls locked until complete</span>"
        "</div>",
        unsafe_allow_html=True
    )
    run_simulation_clicked = False
else:
    run_simulation_clicked = st.sidebar.button(
        "🚀 Run Simulation",
        type="primary",
        use_container_width=True,
        key='run_sim_sidebar_btn',
    )

_algo_map = {
    "Hybrid (Greedy + Monte Carlo)": "hybrid",
    "Monte Carlo Only": "monte_carlo",
    "Greedy + Local Search": "greedy",
}

# ─── DATA PATH ───────────────────────────────────────────────────────────────
DATA_PATH = os.path.join(os.path.dirname(__file__), 'data', 'D1_Port_Pair_Matrix_Advantis.xlsx')

# ── Post-simulation success banner ───────────────────────────────────────────
if st.session_state.get('sim_done') and 'analysis' in st.session_state:
    _res_done = st.session_state['results']
    _el_done  = st.session_state.get('elapsed', 0)
    _its_done = len(_res_done)
    _ips_done = _its_done / max(_el_done, 0.001)
    st.success(
        f"✅ Simulation complete — {_its_done:,} iterations in "
        f"{_el_done:.1f}s ({_ips_done:.0f} iter/sec)  |  "
        f"Best profit: ${max(r['total_profit'] for r in _res_done):,.0f}  |  "
        f"Navigate to any tab above to explore results."
    )

tabs = st.tabs([
    "🌏 Network & Data",
    "📈 Summary Results",
    "⚙️ What-If Editor",
    "📉 Sensitivity",
    "🗺️ Voyage Analysis",
    "🚢 Voyage Journey",
])

# ─── TAB 1: NETWORK & DATA (merged) ──────────────────────────────────────────
with tabs[0]:
    if os.path.exists(DATA_PATH):
        with st.spinner("Loading data..."):
            intra, ports, dist_matrix, legs = load_data(DATA_PATH)

        # ── Top metrics row ───────────────────────────────────────────────
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Ports in Network",   len(ports))
        m2.metric("Feasible Voyage Legs", f"{len(legs):,}")
        m3.metric("Observed Routes",
                  f"{len(legs[legs['status']=='observed']):,}")
        m4.metric("Plausible Routes",
                  f"{len(legs[legs['status']=='plausible']):,}")

        st.markdown("---")

        # ── Side by side: Map LEFT, Data RIGHT ───────────────────────────
        col_map, col_data = st.columns([3, 2], gap="large")

        with col_map:
            st.markdown("### Port Network Map")
            fig_net = px.scatter_map(
                ports, lat='Lat', lon='Lon', size='Total_Vol',
                color='Country', hover_name='Port',
                hover_data={
                    'Load_Vol': ':.0f',
                    'Disch_Vol': ':.0f',
                    'Total_Vol': ':.0f'
                },
                map_style="carto-positron",
                zoom=3, center={"lat": 5, "lon": 115},
                size_max=30,
            )
            fig_net.update_layout(height=500, margin=dict(l=0,r=0,t=0,b=0))
            st.plotly_chart(fig_net, use_container_width=True)

            with st.expander("Distance Matrix Heatmap (Top 20 Ports)"):
                top20_idx   = ports.head(20).index.tolist()
                top20_names = ports.head(20)['Port'].tolist()
                sub_dist    = dist_matrix[np.ix_(top20_idx, top20_idx)]
                fig_dist = px.imshow(
                    sub_dist, x=top20_names, y=top20_names,
                    color_continuous_scale='Blues',
                    title="Sea Distance (NM) — Top 20 Ports"
                )
                fig_dist.update_layout(height=500)
                st.plotly_chart(fig_dist, use_container_width=True)

        with col_data:
            st.markdown("### Port Database")
            disp = ports[[
                'Port_ID', 'Port', 'Country',
                'Load_Vol', 'Disch_Vol', 'Total_Vol', 'Lat', 'Lon'
            ]].copy()
            disp['Load_Vol']  = (disp['Load_Vol']  / 1e6).round(1)
            disp['Disch_Vol'] = (disp['Disch_Vol'] / 1e6).round(1)
            disp['Total_Vol'] = (disp['Total_Vol'] / 1e6).round(1)
            disp.columns = [
                'ID', 'Port', 'Country',
                'Load (M MT)', 'Disch (M MT)', 'Total (M MT)', 'Lat', 'Lon'
            ]
            st.dataframe(disp, use_container_width=True, height=280)

            st.markdown("### Legs by Commodity")
            comm_dist = legs.groupby('commodity').size().sort_values(ascending=True)
            fig_comm = px.bar(
                x=comm_dist.values, y=comm_dist.index,
                orientation='h',
                labels={'x': 'Count', 'y': 'Commodity'},
                color=comm_dist.values,
                color_continuous_scale='Blues',
            )
            fig_comm.update_layout(
                height=380,
                showlegend=False,
                coloraxis_showscale=False,
                margin=dict(l=0, r=0, t=0, b=0),
            )
            st.plotly_chart(fig_comm, use_container_width=True)

            st.markdown("### Legs by Origin Country")
            country_dist = legs.groupby('origin_country').size().sort_values(ascending=True)
            fig_cntry = px.bar(
                x=country_dist.values, y=country_dist.index,
                orientation='h',
                labels={'x': 'Count', 'y': 'Country'},
                color=country_dist.values,
                color_continuous_scale='Teal',
            )
            fig_cntry.update_layout(
                height=260,
                showlegend=False,
                coloraxis_showscale=False,
                margin=dict(l=0, r=0, t=0, b=0),
            )
            st.plotly_chart(fig_cntry, use_container_width=True)

        # ── Network Graph Analysis (full width below) ─────────────────────
        if HAS_NX:
            st.markdown("---")
            st.markdown("### Network Graph Analysis")
            with st.spinner("Building voyage graph..."):
                vessel_for_graph = VesselConfig(
                    dwt=dwt, dwcc=dwcc,
                    speed_laden_knots=speed_laden,
                    speed_ballast_knots=speed_ballast,
                    charter_hire_day=charter_hire,
                    lsfo_price_mt=lsfo_price,
                    mgo_price_mt=mgo_price,
                    insurance_annual=insurance_annual,
                    brokerage_pct=brokerage_pct_ui / 100.0,
                    operating_days_year=operating_days,
                )
                G           = build_voyage_graph(legs, vessel_for_graph)
                centrality  = compute_port_centrality(G)
                communities = find_communities(G)

            if G is not None:
                cent_df = pd.DataFrame([
                    {
                        'Port':       p,
                        'Centrality': round(c, 4),
                        'Community':  communities.get(p, 0)
                    }
                    for p, c in sorted(
                        centrality.items(), key=lambda x: -x[1]
                    )[:30]
                ])
                gc1, gc2 = st.columns(2)
                with gc1:
                    st.markdown("**Top 30 Hub Ports by Betweenness Centrality**")
                    st.dataframe(cent_df, use_container_width=True, height=500)
                with gc2:
                    fig_cent = px.bar(
                        cent_df, x='Centrality', y='Port',
                        orientation='h', color='Community',
                        title="Port Network Centrality",
                        labels={
                            'Centrality': 'Betweenness Centrality',
                            'Port': ''
                        }
                    )
                    fig_cent.update_layout(
                        height=600,
                        yaxis={'categoryorder': 'total ascending'}
                    )
                    st.plotly_chart(fig_cent, use_container_width=True)

                n_communities = len(set(communities.values()))
                st.info(
                    f"Detected **{n_communities} route communities** via "
                    f"greedy modularity detection. "
                    f"Graph: {G.number_of_nodes()} nodes, "
                    f"{G.number_of_edges()} edges."
                )
        else:
            st.warning(
                "Install networkx (`pip install networkx`) "
                "to enable graph analysis."
            )
    else:
        st.error(f"Data file not found: {DATA_PATH}")

# ─── TAB 4: SUMMARY RESULTS ──────────────────────────────────────────────────
with tabs[1]:
    if 'analysis' in st.session_state:
        analysis = st.session_state['analysis']
        s = analysis['summary']

        # ── Hero KPI cards ────────────────────────────────────────────────────
        st.markdown("""
<style>
.kpi-row { display:flex; gap:12px; margin-bottom:1.5rem; flex-wrap:wrap; }
.kpi-card {
    flex:1; min-width:140px; padding:16px 20px;
    background:linear-gradient(135deg,#1a3a5c 0%,#2d5986 100%);
    border-radius:12px; color:white;
}
.kpi-val  { font-size:1.7rem; font-weight:700; letter-spacing:-0.5px; }
.kpi-lbl  { font-size:0.78rem; opacity:0.75; margin-top:4px; }
.kpi-del  { font-size:0.82rem; margin-top:6px; }
.kpi-pos  { color:#86efac; }
.kpi-neg  { color:#fca5a5; }
</style>
""", unsafe_allow_html=True)

        results = st.session_state['results']
        pct_above_market = (
            sum(1 for r in results if r['avg_tce'] >= 8500) / max(len(results), 1) * 100
        )
        st.markdown(f"""
<div class="kpi-row">
  <div class="kpi-card">
    <div class="kpi-val">${s['mean_profit']:,.0f}</div>
    <div class="kpi-lbl">Mean Annual Profit</div>
    <div class="kpi-del kpi-{'pos' if s['mean_profit']>0 else 'neg'}">P90: ${s['p90_profit']:,.0f}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val">${s['median_tce']:,.0f}</div>
    <div class="kpi-lbl">Median TCE ($/day)</div>
    <div class="kpi-del">Market avg ~$8,500</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val">{s['profitable_pct']:.1f}%</div>
    <div class="kpi-lbl">Programmes Profitable</div>
    <div class="kpi-del">{pct_above_market:.1f}% above market TCE</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val">{s['mean_utilisation']:.1f}%</div>
    <div class="kpi-lbl">Mean Utilisation</div>
    <div class="kpi-del">{s['mean_voyages']:.1f} voyages/year avg</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-val">${s['p10_profit']:,.0f}</div>
    <div class="kpi-lbl">Downside P10 Profit</div>
    <div class="kpi-del">Worst 10% scenario</div>
  </div>
</div>
""", unsafe_allow_html=True)

        st.markdown("### Profit Distribution")
        profits = [r['total_profit'] for r in results]
        tces    = [r['avg_tce'] for r in results]

        MARKET_TCE_BENCHMARKS = {
            'Break-even':        0,
            'Market floor':   5500,
            'SEA market avg': 8500,
            'Strong market': 12000,
            'Excellent':     16000,
        }

        col_dist1, col_dist2 = st.columns(2)
        with col_dist1:
            fig_profit = go.Figure()
            fig_profit.add_trace(go.Histogram(
                x=profits, nbinsx=80, marker_color='#667eea',
                name='Annual Profit', opacity=0.8
            ))
            fig_profit.add_vline(
                x=0, line_dash="dash", line_color="red",
                annotation_text="Break-even", annotation_position="top right"
            )
            fig_profit.add_vline(
                x=s['mean_profit'], line_dash="dot", line_color="#22c55e",
                annotation_text=f"Mean ${s['mean_profit']:,.0f}",
                annotation_position="top left"
            )
            fig_profit.update_layout(
                title="Annual Profit Distribution",
                xaxis_title="USD", height=380
            )
            st.plotly_chart(fig_profit, use_container_width=True)

        with col_dist2:
            fig_tce = go.Figure()
            fig_tce.add_trace(go.Histogram(
                x=tces, nbinsx=80, marker_color='#764ba2',
                name='TCE', opacity=0.8
            ))
            for label, value in MARKET_TCE_BENCHMARKS.items():
                color = 'red' if value == 0 else '#f59e0b' if value <= 8500 else '#22c55e'
                fig_tce.add_vline(
                    x=value, line_dash="dash", line_color=color,
                    annotation_text=f"{label} (${value:,})",
                    annotation_position="top right",
                    annotation_font_size=10,
                )
            fig_tce.update_layout(
                title="TCE Distribution vs Market Benchmarks",
                xaxis_title="USD/day", height=380
            )
            st.plotly_chart(fig_tce, use_container_width=True)

        mean_tce = s['median_tce']
        if mean_tce >= 12000:
            tce_verdict = "🟢 Excellent — well above SEA market average"
        elif mean_tce >= 8500:
            tce_verdict = "🟡 Good — at or above SEA market average"
        elif mean_tce >= 5500:
            tce_verdict = "🟠 Below average — but covering costs"
        else:
            tce_verdict = "🔴 Poor — below typical market floor"
        st.info(
            f"**Median TCE: ${mean_tce:,.0f}/day** — {tce_verdict}  |  "
            f"SEA ~20k DWT handy bulk market reference: $8,500–$12,000/day"
        )

        col1, col2 = st.columns(2)
        with col1:
            port_df = pd.DataFrame(analysis['port_ranking'][:30])
            if len(port_df) > 0:
                fig = px.bar(port_df, x='frequency_pct', y='port', orientation='h',
                             title="Port Frequency in Top 10% Programmes",
                             labels={'frequency_pct': 'Appearance %', 'port': 'Port'})
                fig.update_layout(height=700, yaxis={'categoryorder': 'total ascending'})
                st.plotly_chart(fig, use_container_width=True)
        with col2:
            comm_df = pd.DataFrame(analysis['commodity_ranking'])
            if len(comm_df) > 0:
                fig = px.bar(comm_df, x='avg_profit_per_leg', y='commodity', orientation='h',
                             title="Average Profit per Voyage Leg by Commodity",
                             labels={'avg_profit_per_leg': 'Avg Profit/Leg (USD)', 'commodity': 'Commodity'})
                fig.update_layout(height=700, yaxis={'categoryorder': 'total ascending'})
                st.plotly_chart(fig, use_container_width=True)

        # Radar chart — commodity performance scorecard
        st.markdown("#### Commodity Performance Scorecard")
        st.caption(
            "Each axis scored 0–10 relative to other commodities. "
            "Larger area = better overall performer."
        )

        comm_df_full = pd.DataFrame(analysis['commodity_ranking'])
        if len(comm_df_full) > 0:
            top_comms = comm_df_full.head(6)

            def normalise(series):
                mn, mx = series.min(), series.max()
                if mx == mn:
                    return pd.Series([5.0] * len(series), index=series.index)
                return (series - mn) / (mx - mn) * 10

            scores = pd.DataFrame()
            scores['commodity']     = top_comms['commodity'].values
            scores['Profitability'] = normalise(top_comms['avg_profit_per_leg']).values
            scores['Frequency']     = normalise(top_comms['frequency']).values
            scores['Total Revenue'] = normalise(top_comms['total_revenue']).values
            scores['Total Profit']  = normalise(top_comms['total_profit']).values
            scores['Consistency']   = normalise(
                top_comms['avg_profit_per_leg'] / (
                    top_comms['total_profit'].abs() /
                    top_comms['frequency'].clip(lower=1)
                ).clip(lower=0.1)).values

            categories = ['Profitability', 'Frequency',
                          'Total Revenue', 'Total Profit', 'Consistency']

            RADAR_COLORS = [
                '#667eea', '#22c55e', '#f59e0b',
                '#ef4444', '#06b6d4', '#8b5cf6'
            ]

            fig_radar = go.Figure()
            for i, row in scores.iterrows():
                vals = [row[c] for c in categories]
                vals_closed = vals + [vals[0]]
                cats_closed = categories + [categories[0]]
                fig_radar.add_trace(go.Scatterpolar(
                    r=vals_closed,
                    theta=cats_closed,
                    fill='toself',
                    name=row['commodity'],
                    line=dict(color=RADAR_COLORS[i % len(RADAR_COLORS)], width=2),
                    fillcolor=RADAR_COLORS[i % len(RADAR_COLORS)],
                    opacity=0.15,
                    hovertemplate=(
                        f"<b>{row['commodity']}</b><br>"
                        + "<br>".join(
                            f"{c}: {row[c]:.1f}/10" for c in categories
                        )
                        + "<extra></extra>"
                    ),
                ))

            fig_radar.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        range=[0, 10],
                        tickfont=dict(size=9),
                        gridcolor='rgba(0,0,0,0.1)',
                    ),
                    angularaxis=dict(
                        tickfont=dict(size=11, color='#1a3a5c'),
                        gridcolor='rgba(0,0,0,0.1)',
                    ),
                    bgcolor='rgba(0,0,0,0)',
                ),
                showlegend=True,
                legend=dict(
                    orientation='v',
                    x=1.05, y=0.5,
                    font=dict(size=11),
                ),
                title=dict(
                    text="Top 6 Commodities — Multi-Dimension Performance",
                    font=dict(size=13, color='#1a3a5c'),
                ),
                height=480,
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                margin=dict(l=60, r=160, t=60, b=40),
            )
            st.plotly_chart(fig_radar, use_container_width=True)

        route_df = pd.DataFrame(analysis['route_ranking'][:25])
        if len(route_df) > 0:
            fig = px.bar(route_df, x='avg_profit', y='route', orientation='h',
                         title="Top 25 Routes by Average Profit",
                         labels={'avg_profit': 'Avg Profit (USD)', 'route': 'Route'},
                         color='frequency', color_continuous_scale='Viridis')
            fig.update_layout(height=700, yaxis={'categoryorder': 'total ascending'})
            st.plotly_chart(fig, use_container_width=True)

        net_df = pd.DataFrame(analysis['network_size'])
        if len(net_df) > 0:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=net_df['n_ports'], y=net_df['mean_profit'],
                                 name='Mean Profit', marker_color='#667eea'), secondary_y=False)
            fig.add_trace(go.Scatter(x=net_df['n_ports'], y=net_df['mean_tce'],
                                     name='Mean TCE', line=dict(color='red', width=3)), secondary_y=True)
            fig.update_layout(title="Profit & TCE by Ports in Network", xaxis_title="Number of Ports", height=420)
            fig.update_yaxes(title_text="Mean Profit (USD)", secondary_y=False)
            fig.update_yaxes(title_text="Mean TCE (USD/day)", secondary_y=True)
            st.plotly_chart(fig, use_container_width=True)

        phase_df = pd.DataFrame(analysis['phase_comparison'])
        if len(phase_df) > 0:
            phase_label = {0: 'Greedy', 1: 'Exploration', 2: 'Informed', 3: 'Exploitation'}
            phase_df['phase_name'] = phase_df['phase'].map(phase_label).fillna('Unknown')
            fig = px.bar(phase_df, x='phase_name', y=['mean_profit', 'median_profit', 'max_profit'],
                         barmode='group', title="Profit by Simulation Phase")
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Run the simulation first.")

# ─── TAB 6: WHAT-IF / VOYAGE EDITOR ─────────────────────────────────────────
with tabs[2]:
    if 'analysis' in st.session_state:
        analysis  = st.session_state['analysis']
        vessel    = st.session_state.get('vessel')
        dm        = st.session_state.get('dist_matrix')
        legs_df   = st.session_state.get('legs_df')
        top_progs = analysis['top_programmes']

        if vessel is None or dm is None or legs_df is None:
            st.warning("Re-run the simulation to enable the What-If editor.")
        else:
            st.markdown("### What-If / Voyage Editor")
            st.info(
                "Modify any voyage in a programme and all downstream voyages automatically "
                "cascade-recalculate (ballast distances, costs, cumulative profit)."
            )

            prog_sel = st.selectbox("Load programme", range(1, len(top_progs) + 1),
                                    key='whatif_prog')
            prog = top_progs[prog_sel - 1]

            # Build fast library once
            if 'fast_lib' not in st.session_state:
                from modules.simulation_engine import _ensure_port_cost_columns
                legs_df_v2 = _ensure_port_cost_columns(legs_df)
                st.session_state['fast_lib'] = FastLegLibrary(legs_df_v2, vessel)

            fast_lib = st.session_state['fast_lib']

            # Reconstruct VoyageLeg objects from dicts
            from modules.simulation_engine import VoyageLeg, cascade_recalculate_legs

            def dict_to_leg(d):
                leg = VoyageLeg(
                    origin_id=int(d.get('origin_id', 0)),
                    dest_id=int(d.get('dest_id', 0)),
                    origin_port=d['origin_port'],
                    dest_port=d['dest_port'],
                    ballast_from_port=d.get('ballast_from_port', ''),
                    commodity=d['commodity'],
                    category=d.get('category', ''),
                    cargo_mt=d.get('cargo_mt', vessel.dwcc),
                    freight_rate=d.get('freight_rate', 0),
                    ballast_nm=d.get('ballast_nm', d.get('ballast_distance_nm', 0)),
                    laden_nm=d.get('laden_nm', d.get('distance_nm', 0)),
                    gross_freight=d.get('gross_freight', d.get('revenue', 0)),
                    brokerage=d.get('brokerage', 0),
                    net_income=d.get('net_income', 0),
                    maneuver_days=d.get('maneuver_days', 0.503333),
                    loading_days=d.get('loading_days', d.get('load_days', 0)),
                    discharge_days=d.get('discharge_days', d.get('disch_days', 0)),
                    port_days=d.get('port_days', 0),
                    idle_days=d.get('idle_days', 0),
                    ballast_days=d.get('ballast_days', 0),
                    laden_days=d.get('laden_days', 0),
                    total_days=d.get('total_days', 0),
                    lsfo_mt=d.get('lsfo_mt', 0),
                    mgo_mt=d.get('mgo_mt', 0),
                    lsfo_cost=d.get('lsfo_cost', 0),
                    mgo_cost=d.get('mgo_cost', 0),
                    bunker_cost=d.get('bunker_cost', 0),
                    charter_hire=d.get('charter_hire', d.get('charter_hire_cost', 0)),
                    port_costs=d.get('port_costs', 0),
                    load_port_nav=d.get('load_port_nav', 0),
                    load_port_steve=d.get('load_port_steve', 0),
                    disch_port_nav=d.get('disch_port_nav', 0),
                    disch_port_steve=d.get('disch_port_steve', 0),
                    insurance=d.get('insurance', 0),
                    other_costs=d.get('other_costs', 1000),
                    total_expenses=d.get('total_expenses', d.get('total_cost', 0)),
                    profit_loss=d.get('profit_loss', d.get('profit', 0)),
                    profit_per_day=d.get('profit_per_day', 0),
                    profit_per_mt=d.get('profit_per_mt', 0),
                    cum_days=d.get('cum_days', 0),
                    cum_profit=d.get('cum_profit', 0),
                )
                return leg

            # Initialise session state for the working copy of legs
            prog_key = f'whatif_legs_{prog_sel}'
            if prog_key not in st.session_state:
                st.session_state[prog_key] = [dict_to_leg(d) for d in prog['legs']]

            working_legs = st.session_state[prog_key]

            # ── Voyage selector and editor ────────────────────────────────────
            st.markdown("#### Select a voyage to modify")
            voy_labels = [
                f"V{i+1}: {l.origin_port} → {l.dest_port} | {l.commodity} | ${l.profit_loss:,.0f}"
                for i, l in enumerate(working_legs)
            ]
            edit_idx = st.selectbox("Voyage to edit", range(len(working_legs)),
                                    format_func=lambda i: voy_labels[i], key='whatif_voy')

            sel_leg = working_legs[edit_idx]

            # Collect available load ports and discharge ports from leg library
            load_ports = sorted(legs_df['origin_port'].unique().tolist())
            disch_ports = sorted(legs_df['dest_port'].unique().tolist())
            commodities = sorted(legs_df['commodity'].unique().tolist())

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                new_load = st.selectbox("New Load Port",
                                        load_ports,
                                        index=load_ports.index(sel_leg.origin_port) if sel_leg.origin_port in load_ports else 0,
                                        key='wi_load')
            with col2:
                new_disch = st.selectbox("New Discharge Port",
                                         disch_ports,
                                         index=disch_ports.index(sel_leg.dest_port) if sel_leg.dest_port in disch_ports else 0,
                                         key='wi_disch')
            with col3:
                new_comm = st.selectbox("Commodity",
                                        commodities,
                                        index=commodities.index(sel_leg.commodity) if sel_leg.commodity in commodities else 0,
                                        key='wi_comm')
            with col4:
                new_rate = st.number_input("Freight Rate ($/MT)",
                                           value=float(round(sel_leg.freight_rate, 2)),
                                           min_value=1.0, max_value=200.0, step=0.5,
                                           key='wi_rate')

            if st.button("Apply Change & Cascade Recalculate", type="primary"):
                # Find matching leg in library
                load_row = legs_df[
                    (legs_df['origin_port'] == new_load) &
                    (legs_df['dest_port']   == new_disch) &
                    (legs_df['commodity']   == new_comm)
                ]
                if load_row.empty:
                    st.error(f"No feasible leg found for {new_load} → {new_disch} carrying {new_comm}.")
                else:
                    row = load_row.iloc[0]
                    new_leg_idx = int(row.name) if int(row.name) < len(fast_lib.origin_ids) else _find_leg_idx(
                        fast_lib, int(row['origin_id']), int(row['dest_id']), new_comm
                    )
                    # Determine ballast NM from previous voyage
                    if edit_idx == 0:
                        ballast_nm = 0.0
                        bfrom = new_load
                    else:
                        prev_dest_id = working_legs[edit_idx - 1].dest_id
                        ballast_nm = float(dm[prev_dest_id, int(row['origin_id'])])
                        bfrom = working_legs[edit_idx - 1].dest_port

                    new_leg = fast_lib.build_voyage_leg(
                        _find_leg_idx(fast_lib, int(row['origin_id']), int(row['dest_id']), new_comm),
                        ballast_nm, bfrom,
                        vessel.lsfo_price_mt, vessel.mgo_price_mt,
                        new_rate,
                    )
                    if new_leg is None:
                        st.error("Failed to compute costs for this leg.")
                    else:
                        working_legs[edit_idx] = new_leg
                        # Cascade from the next voyage
                        working_legs = cascade_recalculate_legs(
                            working_legs, edit_idx + 1, dm, vessel, fast_lib,
                            vessel.lsfo_price_mt, vessel.mgo_price_mt,
                        )
                        st.session_state[prog_key] = working_legs
                        st.success(f"Voyage {edit_idx+1} updated — all downstream voyages recalculated.")
                        st.rerun()

            if st.button("Reset to Original Programme"):
                if prog_key in st.session_state:
                    del st.session_state[prog_key]
                st.rerun()

            # ── Before/After comparison ───────────────────────────────────────
            st.markdown("#### Current Programme — Voyage Schedule")
            wi_rows = []
            for i, leg in enumerate(working_legs):
                orig_leg = prog['legs'][i] if i < len(prog['legs']) else {}
                orig_profit = orig_leg.get('profit_loss', orig_leg.get('profit', 0))
                curr_profit = leg.profit_loss
                delta = curr_profit - orig_profit
                wi_rows.append({
                    'Voy #':       i + 1,
                    'Load Port':   leg.origin_port,
                    'Disch Port':  leg.dest_port,
                    'Commodity':   leg.commodity,
                    'Rate $/MT':   f"${leg.freight_rate:.2f}",
                    'Ballast NM':  f"{leg.ballast_nm:,.0f}",
                    'Laden NM':    f"{leg.laden_nm:,.0f}",
                    'Total Days':  f"{leg.total_days:.2f}",
                    'Profit':      f"${curr_profit:,.0f}",
                    'vs Original': f"{'+'if delta>=0 else ''}{delta:,.0f}",
                    'Cum Profit':  f"${leg.cum_profit:,.0f}",
                })
            st.dataframe(pd.DataFrame(wi_rows), use_container_width=True)

            # Totals comparison
            orig_total = sum(l.get('profit_loss', l.get('profit', 0)) for l in prog['legs'])
            new_total  = sum(l.profit_loss for l in working_legs)
            delta_total = new_total - orig_total

            col1, col2, col3 = st.columns(3)
            col1.metric("Original Annual Profit", f"${orig_total:,.0f}")
            col2.metric("Modified Annual Profit",  f"${new_total:,.0f}")
            col3.metric("Change",                  f"${delta_total:+,.0f}",
                        delta_color="normal" if delta_total >= 0 else "inverse")

            # Cumulative profit before/after chart
            orig_cum, new_cum = [], []
            o_run, n_run = 0.0, 0.0
            n_show = max(len(prog['legs']), len(working_legs))
            for i in range(n_show):
                if i < len(prog['legs']):
                    o_run += prog['legs'][i].get('profit_loss', prog['legs'][i].get('profit', 0))
                if i < len(working_legs):
                    n_run += working_legs[i].profit_loss
                orig_cum.append(o_run)
                new_cum.append(n_run)

            fig = go.Figure()
            fig.add_trace(go.Scatter(x=list(range(1, n_show + 1)), y=orig_cum,
                                     name='Original', line=dict(color='#667eea', width=2, dash='dash')))
            fig.add_trace(go.Scatter(x=list(range(1, n_show + 1)), y=new_cum,
                                     name='Modified', line=dict(color='#22c55e', width=3)))
            fig.add_hline(y=0, line_dash='dash', line_color='red')
            fig.update_layout(
                xaxis_title="Voyage Number",
                yaxis_title="Cumulative Profit (USD)",
                title="Cumulative Profit: Original vs Modified",
                height=420,
            )
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Run the simulation first to use the What-If editor.")

# ─── TAB 7: SENSITIVITY ANALYSIS ─────────────────────────────────────────────
with tabs[3]:
    if 'analysis' not in st.session_state:
        st.info("Run the simulation first.")
    else:
        st.markdown("### Sensitivity Analysis")
        st.markdown(
            "Adjust the sliders to see how changes in market conditions "
            "affect the top programme's profitability — instantly."
        )

        analysis = st.session_state['analysis']
        vessel   = st.session_state.get('vessel')
        top_prog = analysis['top_programmes'][0] if analysis['top_programmes'] else None

        if top_prog is None or vessel is None:
            st.warning("No results available.")
        else:
            legs_base   = top_prog['legs']
            base_profit = top_prog['total_profit']
            base_tce    = top_prog['avg_tce']

            st.markdown("#### Market Scenario Sliders")
            col1, col2, col3 = st.columns(3)
            with col1:
                bunker_delta = st.slider(
                    "Bunker Price Change (%)",
                    min_value=-40, max_value=60, value=0, step=5,
                    help="±% change applied to both LSFO and MGO prices"
                )
            with col2:
                freight_delta = st.slider(
                    "Freight Rate Change (%)",
                    min_value=-40, max_value=40, value=0, step=5,
                    help="±% change applied to all freight rates"
                )
            with col3:
                hire_delta = st.slider(
                    "Charter Hire Change (%)",
                    min_value=-30, max_value=50, value=0, step=5,
                    help="±% change in daily charter hire rate"
                )

            def recalculate_sensitivity(legs, bunker_pct, freight_pct, hire_pct):
                """Recompute programme P&L with adjusted market parameters."""
                b_mult = 1 + bunker_pct / 100
                f_mult = 1 + freight_pct / 100
                h_mult = 1 + hire_pct / 100
                new_profit  = 0.0
                new_revenue = 0.0
                new_bunker  = 0.0
                new_hire    = 0.0
                for leg in legs:
                    gross = leg.get('gross_freight', leg.get('revenue', 0)) * f_mult
                    brok  = gross * 0.0375
                    ni    = gross - brok
                    bunk  = leg.get('bunker_cost', 0) * b_mult
                    hire  = leg.get('charter_hire', leg.get('charter_hire_cost', 0)) * h_mult
                    port  = leg.get('port_costs', 0)
                    ins   = leg.get('insurance', 0)
                    oth   = leg.get('other_costs', 1000)
                    exp   = hire + bunk + port + ins + oth
                    pl    = ni - exp
                    new_profit  += pl
                    new_revenue += gross
                    new_bunker  += bunk
                    new_hire    += hire
                return {
                    'profit': new_profit,
                    'revenue': new_revenue,
                    'bunker': new_bunker,
                    'hire': new_hire,
                }

            result       = recalculate_sensitivity(legs_base, bunker_delta, freight_delta, hire_delta)
            new_profit   = result['profit']
            profit_delta = new_profit - base_profit
            profit_delta_pct = (profit_delta / abs(base_profit)) * 100 if base_profit != 0 else 0
            total_days   = sum(l.get('total_days', 0) for l in legs_base)
            new_tce = (
                result['revenue'] * (1 - 0.0375)
                - result['bunker']
                - sum(l.get('port_costs', 0) for l in legs_base)
                - sum(l.get('insurance', 0) for l in legs_base)
                - sum(l.get('other_costs', 1000) for l in legs_base)
            ) / max(total_days, 1)

            st.markdown("#### Adjusted Programme KPIs")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Base Profit",     f"${base_profit:,.0f}")
            c2.metric("Adjusted Profit", f"${new_profit:,.0f}",
                      delta=f"${profit_delta:+,.0f} ({profit_delta_pct:+.1f}%)",
                      delta_color="normal" if profit_delta >= 0 else "inverse")
            c3.metric("Adjusted TCE",    f"${new_tce:,.0f}/day",
                      delta=f"${new_tce - base_tce:+,.0f}/day",
                      delta_color="normal" if new_tce >= base_tce else "inverse")
            c4.metric("Break-even?",
                      "✅ Profitable" if new_profit > 0 else "❌ Loss-making")

            # Tornado chart
            st.markdown("#### Tornado Chart — Single Variable Impact")
            test_ranges = {
                'Bunker +40%':              recalculate_sensitivity(legs_base,  40,   0,   0)['profit'],
                'Bunker -40%':              recalculate_sensitivity(legs_base, -40,   0,   0)['profit'],
                'Freight +20%':             recalculate_sensitivity(legs_base,   0,  20,   0)['profit'],
                'Freight -20%':             recalculate_sensitivity(legs_base,   0, -20,   0)['profit'],
                'Hire +30%':                recalculate_sensitivity(legs_base,   0,   0,  30)['profit'],
                'Hire -30%':                recalculate_sensitivity(legs_base,   0,   0, -30)['profit'],
                'Bunker +20%, Freight -10%':recalculate_sensitivity(legs_base,  20, -10,   0)['profit'],
            }
            tornado_df = pd.DataFrame([
                {
                    'Scenario': k,
                    'Profit': v,
                    'Delta': v - base_profit,
                    'Color': '#22c55e' if v >= base_profit else '#ef4444',
                }
                for k, v in test_ranges.items()
            ]).sort_values('Delta')

            fig_tornado = go.Figure()
            fig_tornado.add_trace(go.Bar(
                y=tornado_df['Scenario'],
                x=tornado_df['Delta'],
                orientation='h',
                marker_color=tornado_df['Color'].tolist(),
                text=[f"${v:,.0f}" for v in tornado_df['Profit']],
                textposition='outside',
            ))
            fig_tornado.add_vline(x=0, line_color='#1a3a5c', line_width=2)
            fig_tornado.update_layout(
                title="Profit impact of single-variable changes vs base case",
                xaxis_title="Change in Annual Profit (USD)",
                height=420,
                yaxis=dict(tickfont=dict(size=11)),
            )
            st.plotly_chart(fig_tornado, use_container_width=True)

            # Scenario comparison table
            st.markdown("#### Scenario Comparison Table")
            scenarios = {
                'Base case':             (0,   0,   0),
                'Bunker spike +30%':     (30,  0,   0),
                'Freight market -15%':   (0,  -15,  0),
                'Hire +20%':             (0,   0,  20),
                'Bull market (fr +20%)': (0,  20,   0),
                'Bear market':           (20, -20,  10),
                'Perfect storm':         (40, -20,  20),
                'Your current sliders':  (bunker_delta, freight_delta, hire_delta),
            }
            scen_rows = []
            for name, (b, f, h) in scenarios.items():
                r = recalculate_sensitivity(legs_base, b, f, h)
                scen_rows.append({
                    'Scenario':      name,
                    'Bunker Δ':      f"{b:+d}%",
                    'Freight Δ':     f"{f:+d}%",
                    'Hire Δ':        f"{h:+d}%",
                    'Annual Profit': f"${r['profit']:,.0f}",
                    'vs Base':       f"${r['profit'] - base_profit:+,.0f}",
                    'Profitable':    '✅' if r['profit'] > 0 else '❌',
                })
            st.dataframe(
                pd.DataFrame(scen_rows),
                use_container_width=True,
                hide_index=True,
                height=320,
            )

# ─── SIMULATION RUN (outside tabs — triggered from sidebar) ──────────────────
if run_simulation_clicked and os.path.exists(DATA_PATH):
    st.session_state['sim_running'] = True
    st.session_state['sim_done']    = False
    intra, ports, dist_matrix, legs = load_data(DATA_PATH)

    vessel = VesselConfig(
        dwt=dwt, dwcc=dwcc,
        speed_laden_knots=speed_laden,
        speed_ballast_knots=speed_ballast,
        charter_hire_day=charter_hire,
        lsfo_price_mt=lsfo_price,
        mgo_price_mt=mgo_price,
        insurance_annual=insurance_annual,
        brokerage_pct=brokerage_pct_ui / 100.0,
        operating_days_year=operating_days,
        fuel_laden_mt_day=13.5,
        fuel_ballast_mt_day=13.5,
        bunker_price_mt=lsfo_price,
    )
    sim_config = SimConfig(
        n_iterations=n_iterations,
        algorithm=_algo_map[algo_choice],
        freight_volatility=freight_vol,
        bunker_volatility=bunker_vol,
    )

    # ── Modal overlay container ───────────────────────────────────────────
    overlay = st.empty()
    with overlay.container():
        st.markdown(
            "<h2 style='text-align:center;color:#1a3a5c;"
            "margin-bottom:0.25rem'>🚀 Simulation Running</h2>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<p style='text-align:center;color:#64748b;margin-top:0'>"
            f"Running {n_iterations:,} iterations — "
            f"{algo_choice}</p>",
            unsafe_allow_html=True
        )
        st.markdown("---")

        prog_col1, prog_col2, prog_col3, prog_col4 = st.columns(4)
        live_phase      = prog_col1.empty()
        live_best       = prog_col2.empty()
        live_tce        = prog_col3.empty()
        live_speed      = prog_col4.empty()
        progress_bar    = st.progress(0)
        live_status     = st.empty()
        live_chart_slot = st.empty()

        live_phase.markdown(
            "<div style='background:#f1f5f9;border-left:4px solid #94a3b8;"
            "padding:12px 16px;border-radius:8px'>"
            "<div style='font-size:0.75rem;color:#64748b'>Current Phase</div>"
            "<div style='font-size:1rem;font-weight:700;color:#475569;"
            "margin-top:4px'>⏳ Starting...</div>"
            "</div>",
            unsafe_allow_html=True
        )
        live_best.markdown(
            "<div style='background:#f1f5f9;border-left:4px solid #94a3b8;"
            "padding:12px 16px;border-radius:8px'>"
            "<div style='font-size:0.75rem;color:#64748b'>Best Profit Found</div>"
            "<div style='font-size:1.4rem;font-weight:700;color:#475569;"
            "margin-top:4px'>—</div>"
            "</div>",
            unsafe_allow_html=True
        )
        live_tce.markdown(
            "<div style='background:#f1f5f9;border-left:4px solid #94a3b8;"
            "padding:12px 16px;border-radius:8px'>"
            "<div style='font-size:0.75rem;color:#64748b'>Best TCE</div>"
            "<div style='font-size:1.4rem;font-weight:700;color:#475569;"
            "margin-top:4px'>—</div>"
            "</div>",
            unsafe_allow_html=True
        )
        live_speed.markdown(
            "<div style='background:#f1f5f9;border-left:4px solid #94a3b8;"
            "padding:12px 16px;border-radius:8px'>"
            "<div style='font-size:0.75rem;color:#64748b'>Speed</div>"
            "<div style='font-size:1.4rem;font-weight:700;color:#475569;"
            "margin-top:4px'>—</div>"
            "</div>",
            unsafe_allow_html=True
        )

        start_time           = time.time()
        _best_profit_history = []
        _best_profit_so_far  = [float('-inf')]
        _best_tce_so_far     = [0.0]
        _iter_count          = [0]

        PHASE_COLORS = {
            'Greedy + Local Search':             ('#1a3a5c', '🔵'),
            'Phase 1: Pure Exploration':         ('#f59e0b', '🟡'),
            'Phase 2: Informed Exploration':     ('#667eea', '🟣'),
            'Phase 3: Intensive Exploitation':   ('#22c55e', '🟢'),
            'Complete':                          ('#22c55e', '✅'),
        }

        def progress_callback(phase_name, current, total):
            pct       = current / max(1, total)
            elapsed   = time.time() - start_time
            iters_sec = current / max(elapsed, 0.001)
            eta       = (total - current) / max(iters_sec, 0.001)
            progress_bar.progress(pct)
            _iter_count[0] = current
            color, icon = PHASE_COLORS.get(phase_name, ('#667eea', '⚙️'))
            live_phase.markdown(
                f"<div style='background:linear-gradient("
                f"135deg,{color}22,{color}44);"
                f"border-left:4px solid {color};"
                f"padding:12px 16px;border-radius:8px'>"
                f"<div style='font-size:0.75rem;color:#64748b'>"
                f"Current Phase</div>"
                f"<div style='font-size:1rem;font-weight:700;"
                f"color:{color};margin-top:4px'>{icon} "
                f"{phase_name.replace('Phase 1: ','').replace('Phase 2: ','').replace('Phase 3: ','')}"
                f"</div>"
                f"<div style='font-size:0.78rem;color:#64748b;"
                f"margin-top:4px'>{current:,}/{total:,} "
                f"({pct*100:.0f}%)</div>"
                f"</div>",
                unsafe_allow_html=True
            )
            live_speed.markdown(
                f"<div style='background:linear-gradient("
                f"135deg,#f8fafc,#e2e8f0);"
                f"padding:12px 16px;border-radius:8px'>"
                f"<div style='font-size:0.75rem;color:#64748b'>"
                f"Speed</div>"
                f"<div style='font-size:1.4rem;font-weight:700;"
                f"color:#1a3a5c;margin-top:4px'>{iters_sec:.0f}</div>"
                f"<div style='font-size:0.78rem;color:#64748b'>"
                f"iter/sec  |  ETA {eta:.0f}s</div>"
                f"</div>",
                unsafe_allow_html=True
            )
            live_status.markdown(
                f"<div style='text-align:center;color:#64748b;"
                f"font-size:0.82rem;padding:4px 0'>"
                f"Elapsed: {elapsed:.0f}s  |  "
                f"Iterations: {current:,}  |  "
                f"Remaining: {total - current:,}"
                f"</div>",
                unsafe_allow_html=True
            )

        def update_best_metrics(results_so_far):
            if not results_so_far:
                return
            best = max(results_so_far, key=lambda r: r['total_profit'])
            bp   = best['total_profit']
            bt   = best['avg_tce']
            if bp > _best_profit_so_far[0]:
                _best_profit_so_far[0] = bp
                _best_tce_so_far[0]    = bt
                _best_profit_history.append(bp)
                live_best.markdown(
                    f"<div style='background:linear-gradient("
                    f"135deg,#22c55e22,#22c55e44);"
                    f"border-left:4px solid #22c55e;"
                    f"padding:12px 16px;border-radius:8px'>"
                    f"<div style='font-size:0.75rem;color:#64748b'>"
                    f"Best Profit Found</div>"
                    f"<div style='font-size:1.4rem;font-weight:700;"
                    f"color:#166534;margin-top:4px'>${bp:,.0f}</div>"
                    f"<div style='font-size:0.78rem;color:#64748b;"
                    f"margin-top:4px'>After {_iter_count[0]:,} "
                    f"iterations</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                live_tce.markdown(
                    f"<div style='background:linear-gradient("
                    f"135deg,#667eea22,#667eea44);"
                    f"border-left:4px solid #667eea;"
                    f"padding:12px 16px;border-radius:8px'>"
                    f"<div style='font-size:0.75rem;color:#64748b'>"
                    f"Best TCE</div>"
                    f"<div style='font-size:1.4rem;font-weight:700;"
                    f"color:#1a3a5c;margin-top:4px'>${bt:,.0f}/day</div>"
                    f"<div style='font-size:0.78rem;color:#64748b;"
                    f"margin-top:4px'>"
                    f"{'🟢 Above market' if bt >= 8500 else '🟡 Below market avg'}"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                if len(_best_profit_history) >= 2:
                    fig_spark = go.Figure()
                    fig_spark.add_trace(go.Scatter(
                        x=list(range(len(_best_profit_history))),
                        y=_best_profit_history,
                        mode='lines',
                        fill='tozeroy',
                        line=dict(color='#22c55e', width=2),
                        fillcolor='rgba(34,197,94,0.15)',
                        showlegend=False,
                    ))
                    fig_spark.update_layout(
                        title=dict(
                            text="Best profit improving over simulation",
                            font=dict(size=11, color='#64748b')
                        ),
                        xaxis=dict(visible=False),
                        yaxis=dict(
                            tickformat='$,.0f',
                            tickfont=dict(size=9),
                            gridcolor='rgba(0,0,0,0.05)',
                        ),
                        height=140,
                        margin=dict(l=60, r=10, t=28, b=10),
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                    )
                    live_chart_slot.plotly_chart(
                        fig_spark, use_container_width=True
                    )

        results = run_full_simulation(
            legs, dist_matrix, vessel, sim_config, progress_callback,
            port_charge_vol=port_charge_vol,
            congestion_vol=congestion_vol,
            use_stevedoring=not fio_cargo,
        )
        elapsed   = time.time() - start_time
        iters_sec = len(results) / max(elapsed, 0.001)

        update_best_metrics(results)

        live_status.markdown(
            f"<div style='text-align:center;background:#dcfce7;"
            f"color:#166534;font-weight:700;padding:10px;"
            f"border-radius:8px;font-size:1rem'>"
            f"✅ Complete — {len(results):,} iterations in "
            f"{elapsed:.1f}s ({iters_sec:.0f} iter/sec)"
            f"</div>",
            unsafe_allow_html=True
        )

        analysis = analyse_results(results, ports)
        st.session_state['results']     = results
        st.session_state['analysis']    = analysis
        st.session_state['elapsed']     = elapsed
        st.session_state['vessel']      = vessel
        st.session_state['dist_matrix'] = dist_matrix
        st.session_state['legs_df']     = legs
        st.session_state['sim_running'] = False
        st.session_state['sim_done']    = True

    overlay.empty()
    # Rerun so all tabs render with the newly stored results
    st.rerun()

# ─── TAB 6: VOYAGE ANALYSIS ──────────────────────────────────────────────────
with tabs[4]:
    if 'analysis' not in st.session_state:
        st.info(
            "Run the simulation first — click 🚀 Run Simulation "
            "in the sidebar to generate voyage programmes."
        )
    else:
        import math as _math

        analysis  = st.session_state['analysis']
        vessel    = st.session_state.get('vessel')
        top_progs = analysis['top_programmes']

        intra_va, ports_va, dist_matrix_va, legs_va = load_data(DATA_PATH)
        port_coords_va  = dict(zip(
            ports_va['Port'], zip(ports_va['Lat'], ports_va['Lon'])
        ))
        port_country_va = dict(zip(ports_va['Port'], ports_va['Country']))

        from data.port_charges import PORT_CHARGES, PORT_CHARGES_DEFAULT
        def _get_meta(pname):
            pc = PORT_CHARGES.get(pname, PORT_CHARGES_DEFAULT)
            return {
                'nav':     pc.get('nav', 10000),
                'cong':    pc.get('cong_mean', 1.0),
                'country': port_country_va.get(pname, '—'),
            }

        def _calc_nm(pa, pb):
            """Haversine NM with maritime routing correction factor."""
            if pa not in port_coords_va or pb not in port_coords_va:
                return 0
            la1, lo1 = port_coords_va[pa]
            la2, lo2 = port_coords_va[pb]
            R = 3440.065
            dlat = _math.radians(la2 - la1)
            dlon = _math.radians(lo2 - lo1)
            a = (_math.sin(dlat/2)**2
                 + _math.cos(_math.radians(la1)) * _math.cos(_math.radians(la2))
                 * _math.sin(dlon/2)**2)
            straight = 2 * R * _math.asin(_math.sqrt(a))
            c1 = port_country_va.get(pa, '')
            c2 = port_country_va.get(pb, '')
            f = (1.45 if c1 == 'Indonesia' and c2 == 'Indonesia'
                 else 1.40 if 'Bangladesh' in [c1, c2]
                 else 1.35 if 'Philippines' in [c1, c2]
                 else 1.25)
            return round(straight * f)

        def _geodesic(lat1, lon1, lat2, lon2, n=80):
            """Build geodesic arc between two coordinates."""
            if lat1 == lat2 and lon1 == lon2:
                return [lat1], [lon1]
            la1 = _math.radians(lat1); lo1 = _math.radians(lon1)
            la2 = _math.radians(lat2); lo2 = _math.radians(lon2)
            d = 2 * _math.asin(_math.sqrt(
                _math.sin((la2-la1)/2)**2
                + _math.cos(la1) * _math.cos(la2)
                * _math.sin((lo2-lo1)/2)**2
            ))
            if d == 0:
                return [lat1, lat2], [lon1, lon2]
            lats, lons = [], []
            for i in range(n + 1):
                f = i / n
                A = _math.sin((1-f)*d) / _math.sin(d)
                B = _math.sin(f*d)     / _math.sin(d)
                x = (A*_math.cos(la1)*_math.cos(lo1)
                     + B*_math.cos(la2)*_math.cos(lo2))
                y = (A*_math.cos(la1)*_math.sin(lo1)
                     + B*_math.cos(la2)*_math.sin(lo2))
                z = A*_math.sin(la1) + B*_math.sin(la2)
                lats.append(_math.degrees(_math.atan2(z, _math.sqrt(x*x + y*y))))
                lons.append(_math.degrees(_math.atan2(y, x)))
            return lats, lons

        # ── Programme selector header ─────────────────────────────────────
        hcol1, hcol2, hcol3 = st.columns([3, 1, 1])
        with hcol1:
            prog_opts = [
                f"#{p['rank']}  —  ${p['total_profit']:,.0f}  —  "
                f"{p['n_voyages']} voyages  —  TCE ${p['avg_tce']:,.0f}/day"
                for p in top_progs
            ]
            sel_prog_idx = st.selectbox(
                "Programme",
                range(len(top_progs)),
                format_func=lambda i: prog_opts[i],
                key='va_prog_sel',
                label_visibility='collapsed',
            )
        prog_va   = top_progs[sel_prog_idx]
        legs_list = prog_va['legs']

        with hcol2:
            st.metric("Annual Profit", f"${prog_va['total_profit']:,.0f}")
        with hcol3:
            try:
                excel_bytes = export_programme_to_excel(
                    prog_va, vessel, prog_va['rank']
                )
                st.download_button(
                    label="📥 Export Excel",
                    data=excel_bytes,
                    file_name=f"voyage_analysis_{prog_va['rank']}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                )
            except Exception:
                pass

        st.markdown("---")

        # ── Selected voyage tracker ───────────────────────────────────────
        va_key = f"va_sel_{sel_prog_idx}"
        if va_key not in st.session_state:
            st.session_state[va_key] = 0
        sel_idx = min(st.session_state[va_key], len(legs_list) - 1)

        # ════════════════════════════════════════════════════════════════
        # ZONE 1 — Left: voyage list + cum chart | Right: config + P&L
        # ════════════════════════════════════════════════════════════════
        z1_left, z1_right = st.columns([2, 3], gap="large")

        # ── LEFT: Compact voyage list ─────────────────────────────────
        with z1_left:
            st.markdown(
                "<div style='font-size:12px;color:#64748b;"
                "margin-bottom:6px'>Click any voyage to explore</div>",
                unsafe_allow_html=True
            )

            for i, leg in enumerate(legs_list):
                pl   = leg.get('profit_loss', leg.get('profit', 0))
                is_s = (i == sel_idx)
                tg_bg = "#dcfce7" if pl >= 0 else "#fee2e2"
                tg_tx = "#166534" if pl >= 0 else "#991b1b"
                sign  = "+" if pl >= 0 else ""

                bc1, bc2 = st.columns([3, 1])
                with bc1:
                    lbl = (
                        f"{'▼ ' if is_s else ''}"
                        f"**{i+1}.** "
                        f"{leg.get('origin_port','?')} → "
                        f"{leg.get('dest_port','?')}"
                    )
                    if st.button(
                        lbl,
                        key=f"va_btn_{sel_prog_idx}_{i}",
                        use_container_width=True,
                        type="secondary",
                    ):
                        st.session_state[va_key] = i
                        st.rerun()
                with bc2:
                    st.markdown(
                        f"<div style='text-align:right;padding-top:4px'>"
                        f"<span style='font-size:11px;background:{tg_bg};"
                        f"color:{tg_tx};padding:2px 7px;border-radius:12px;"
                        f"font-weight:500'>{sign}${pl/1000:.0f}k</span>"
                        f"<div style='font-size:10px;color:#94a3b8;"
                        f"text-align:right;margin-top:1px'>"
                        f"{leg.get('total_days',0):.1f}d</div></div>",
                        unsafe_allow_html=True
                    )

            # Cumulative profit bar chart
            st.markdown(
                "<div style='font-size:11px;color:#94a3b8;"
                "margin-top:10px;margin-bottom:4px'>"
                "Cumulative profit build-up</div>",
                unsafe_allow_html=True
            )
            cum_vals  = [l.get('cum_profit', 0) for l in legs_list]
            pl_vals   = [l.get('profit_loss', l.get('profit', 0))
                         for l in legs_list]
            bar_colors_va = [
                '#22c55e' if v >= 0 else '#ef4444' for v in pl_vals
            ]
            bar_colors_va[sel_idx] = '#f59e0b'

            fig_cum_va = go.Figure()
            fig_cum_va.add_trace(go.Bar(
                x=list(range(1, len(legs_list) + 1)),
                y=cum_vals,
                marker_color=bar_colors_va,
                showlegend=False,
                hovertemplate='Voyage %{x}<br>Cum: $%{y:,.0f}<extra></extra>',
            ))
            fig_cum_va.add_hline(y=0, line_color='#1a3a5c', line_width=1)
            fig_cum_va.add_vline(
                x=sel_idx + 1, line_color='#f59e0b',
                line_width=2, line_dash='dot',
            )
            fig_cum_va.update_layout(
                height=160, margin=dict(l=0, r=0, t=0, b=0),
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(title='Voyage', tickfont=dict(size=9),
                           gridcolor='rgba(0,0,0,0.05)'),
                yaxis=dict(tickformat='$,.0f', tickfont=dict(size=9),
                           gridcolor='rgba(0,0,0,0.05)'),
            )
            st.plotly_chart(fig_cum_va, use_container_width=True)

        # ── RIGHT: Config cards + P&L ─────────────────────────────────
        with z1_right:
            sel_leg = legs_list[sel_idx]
            pl_val  = sel_leg.get('profit_loss', sel_leg.get('profit', 0))
            pl_col_h = "#166534" if pl_val >= 0 else "#991b1b"
            pl_bg_h  = "#dcfce7" if pl_val >= 0 else "#fee2e2"

            st.markdown(
                f"<div style='background:#f8fafc;border:0.5px solid #e2e8f0;"
                f"border-radius:10px;padding:10px 14px;margin-bottom:10px;"
                f"display:flex;align-items:center;gap:12px;flex-wrap:wrap'>"
                f"<span style='font-size:14px;font-weight:600;color:#1a3a5c'>"
                f"Voyage {sel_idx+1} of {len(legs_list)}</span>"
                f"<span style='font-size:13px;color:#475569'>"
                f"{sel_leg.get('origin_port','?')} → "
                f"{sel_leg.get('dest_port','?')}</span>"
                f"<span style='background:#e6f1fb;color:#0c447c;"
                f"font-size:11px;padding:2px 8px;border-radius:10px'>"
                f"{sel_leg.get('commodity','')}</span>"
                f"<span style='margin-left:auto;background:{pl_bg_h};"
                f"color:{pl_col_h};font-size:13px;font-weight:600;"
                f"padding:3px 10px;border-radius:10px'>"
                f"{'+'if pl_val>=0 else ''}${pl_val:,.0f}</span>"
                f"</div>",
                unsafe_allow_html=True
            )

            cc1, cc2, cc3 = st.columns(3)

            # ── Card 1: Ports ──────────────────────────────────────────
            with cc1:
                st.markdown(
                    "<div style='font-size:10px;font-weight:600;color:#64748b;"
                    "text-transform:uppercase;letter-spacing:.06em;"
                    "margin-bottom:6px'>Ports</div>",
                    unsafe_allow_html=True
                )
                all_load  = sorted(legs_va['origin_port'].unique().tolist())
                all_disch = sorted(legs_va['dest_port'].unique().tolist())
                all_comm  = sorted(legs_va['commodity'].unique().tolist())

                cur_load  = sel_leg.get('origin_port',  all_load[0])
                cur_disch = sel_leg.get('dest_port',    all_disch[0])
                cur_comm  = sel_leg.get('commodity',    all_comm[0])

                new_load = st.selectbox(
                    "Load port", all_load,
                    index=(all_load.index(cur_load)
                           if cur_load in all_load else 0),
                    key=f'va_load_{sel_prog_idx}_{sel_idx}',
                )
                new_disch = st.selectbox(
                    "Discharge port", all_disch,
                    index=(all_disch.index(cur_disch)
                           if cur_disch in all_disch else 0),
                    key=f'va_disch_{sel_prog_idx}_{sel_idx}',
                )
                new_comm = st.selectbox(
                    "Commodity", all_comm,
                    index=(all_comm.index(cur_comm)
                           if cur_comm in all_comm else 0),
                    key=f'va_comm_{sel_prog_idx}_{sel_idx}',
                )
                load_meta  = _get_meta(new_load)
                disch_meta = _get_meta(new_disch)
                st.markdown(
                    f"<div style='font-size:10px;color:#64748b;margin-top:4px'>"
                    f"Load nav: <b>${load_meta['nav']:,}</b> · "
                    f"Cong: <b>{load_meta['cong']:.1f}d</b><br>"
                    f"Disch nav: <b>${disch_meta['nav']:,}</b> · "
                    f"Cong: <b>{disch_meta['cong']:.1f}d</b>"
                    f"</div>",
                    unsafe_allow_html=True
                )

            # ── Card 2: Distances & Days ───────────────────────────────
            with cc2:
                st.markdown(
                    "<div style='font-size:10px;font-weight:600;color:#64748b;"
                    "text-transform:uppercase;letter-spacing:.06em;"
                    "margin-bottom:6px'>Distances &amp; days</div>",
                    unsafe_allow_html=True
                )
                auto_laden = _calc_nm(new_load, new_disch)
                _prev_port = (
                    legs_list[sel_idx-1].get('dest_port', 'prev')
                    if sel_idx > 0 else 'start'
                )
                auto_ballast = (
                    _calc_nm(_prev_port, new_load) if sel_idx > 0 else 0
                )
                prev_lbl = (
                    f"from {_prev_port}" if sel_idx > 0 else "first voyage"
                )
                spd_l = vessel.speed_laden_knots   if vessel else 11.0
                spd_b = vessel.speed_ballast_knots if vessel else 11.5

                laden_nm = st.number_input(
                    f"Laden NM (auto: {auto_laden:,})",
                    value=float(auto_laden),
                    min_value=0.0, step=10.0,
                    key=f'va_lnm_{sel_prog_idx}_{sel_idx}_{new_load}_{new_disch}',
                    help="Auto-calculated from port coordinates. Override if needed."
                )
                laden_days = laden_nm / (spd_l * 24)
                st.caption(f"@ {spd_l} kn → **{laden_days:.2f} days**")

                ballast_nm = st.number_input(
                    f"Ballast NM ({prev_lbl})",
                    value=float(auto_ballast),
                    min_value=0.0, step=10.0,
                    key=f'va_bnm_{sel_prog_idx}_{sel_idx}_{_prev_port}_{new_load}',
                )
                ballast_days = (
                    ballast_nm / (spd_b * 24) if ballast_nm > 0 else 0
                )
                st.caption(f"@ {spd_b} kn → **{ballast_days:.2f} days**")

                st.markdown(
                    f"<div style='background:#f0f9ff;border:0.5px solid "
                    f"#bae6fd;border-radius:6px;padding:6px 8px;"
                    f"margin-top:4px;font-size:11px;color:#0369a1'>"
                    f"Total sailing: "
                    f"<b>{laden_days+ballast_days:.2f} days</b></div>",
                    unsafe_allow_html=True
                )

            # ── Card 3: Cargo & Freight ────────────────────────────────
            with cc3:
                st.markdown(
                    "<div style='font-size:10px;font-weight:600;color:#64748b;"
                    "text-transform:uppercase;letter-spacing:.06em;"
                    "margin-bottom:6px'>Cargo &amp; freight</div>",
                    unsafe_allow_html=True
                )
                dwcc_v = vessel.dwcc if vessel else 15000
                new_cargo = st.number_input(
                    "Cargo (MT)",
                    value=float(sel_leg.get('cargo_mt', dwcc_v)),
                    min_value=1000.0, max_value=float(dwcc_v),
                    step=500.0,
                    key=f'va_cargo_{sel_prog_idx}_{sel_idx}_{new_load}_{new_disch}',
                )
                new_rate = st.number_input(
                    "Freight rate ($/MT)",
                    value=float(round(sel_leg.get('freight_rate', 15.0), 2)),
                    min_value=1.0, max_value=200.0, step=0.5,
                    key=f'va_rate_{sel_prog_idx}_{sel_idx}_{new_comm}',
                )
                gross_c = new_cargo * new_rate
                ni_c    = gross_c * (1 - 0.0375)
                st.markdown(
                    f"<div style='background:#f8fafc;border:0.5px solid "
                    f"#e2e8f0;border-radius:6px;padding:6px 8px;margin-top:4px'>"
                    f"<div style='font-size:10px;color:#64748b'>Gross freight</div>"
                    f"<div style='font-size:16px;font-weight:600;color:#1a3a5c'>"
                    f"${gross_c:,.0f}</div>"
                    f"<div style='font-size:10px;color:#64748b;margin-top:2px'>"
                    f"Net income: ${ni_c:,.0f}</div></div>",
                    unsafe_allow_html=True
                )
                st.markdown("<br>", unsafe_allow_html=True)
                recalc_va = st.button(
                    "Recalculate",
                    key=f'va_recalc_{sel_prog_idx}_{sel_idx}',
                    use_container_width=True,
                    type="primary",
                )
                st.button(
                    "Cascade ↗",
                    key=f'va_casc_{sel_prog_idx}_{sel_idx}',
                    use_container_width=True,
                )

            # ── Live P&L computation ──────────────────────────────────
            computed_va = None
            if vessel:
                try:
                    from modules.simulation_engine import cost_voyage_exact
                    from modules.data_processor import (
                        CARGO_RATE, COMMODITY_CATEGORIES
                    )
                    cat_va   = COMMODITY_CATEGORIES.get(new_comm, 'Dry Bulk')
                    rates_va = CARGO_RATE.get(
                        cat_va, {'load': 5000, 'disch': 4000}
                    )
                    lm_va  = _get_meta(new_load)
                    dm_va  = _get_meta(new_disch)
                    computed_va = cost_voyage_exact(
                        cargo_mt=new_cargo,
                        freight_rate=new_rate,
                        laden_nm=laden_nm,
                        ballast_nm=ballast_nm,
                        load_rate_mt_day=rates_va['load'],
                        disch_rate_mt_day=rates_va['disch'],
                        load_port_nav=lm_va['nav'],
                        load_port_steve=0,
                        disch_port_nav=dm_va['nav'],
                        disch_port_steve=0,
                        vessel=vessel,
                        load_cong_days=lm_va['cong'],
                        disch_cong_days=dm_va['cong'],
                    )
                except Exception:
                    computed_va = None

            if recalc_va and computed_va:
                st.success(
                    f"Recalculated — "
                    f"P&L: {'+'if computed_va['profit_loss']>=0 else ''}"
                    f"${computed_va['profit_loss']:,.0f}  |  "
                    f"TCE: ${computed_va['profit_per_day']:,.0f}/day"
                )

            st.markdown("---")

            # ── P&L table + Waterfall chart ───────────────────────────
            pl_col_l, wf_col_r = st.columns(2)

            def _v_va(k, fb=0):
                if computed_va and k in computed_va:
                    return computed_va[k]
                return sel_leg.get(k, fb)

            g_va   = _v_va('gross_freight', sel_leg.get('revenue', 0))
            br_va  = _v_va('brokerage')
            ni_va  = _v_va('net_income')
            ch_va  = _v_va('charter_hire', sel_leg.get('charter_hire_cost', 0))
            lc_va  = _v_va('lsfo_cost')
            mc_va  = _v_va('mgo_cost')
            tb_va  = lc_va + mc_va
            pc_va  = _v_va('port_costs',
                            sel_leg.get('load_port_nav', 0)
                            + sel_leg.get('disch_port_nav', 0))
            ins_va = _v_va('insurance')
            oth_va = _v_va('other_costs', 1000)
            exp_va = _v_va('total_expenses', sel_leg.get('total_cost', 0))
            pl_va  = _v_va('profit_loss',   sel_leg.get('profit', 0))
            ppd_va = _v_va('profit_per_day')
            td_va  = _v_va('total_days',    sel_leg.get('total_days', 0))
            vs_avg_va = ppd_va - prog_va['avg_tce']

            with pl_col_l:
                st.markdown(
                    "<div style='font-size:10px;font-weight:600;color:#64748b;"
                    "text-transform:uppercase;letter-spacing:.06em;"
                    "margin-bottom:8px'>P&amp;L breakdown</div>",
                    unsafe_allow_html=True
                )
                pl_items_va = [
                    ("Gross freight",      g_va,              False, False),
                    ("Brokerage (3.75%)", -br_va,             True,  False),
                    ("Net income",         ni_va,             False, True),
                    ("Charter hire",      -ch_va,             True,  False),
                    ("LSFO bunker",       -lc_va,             True,  False),
                    ("MGO bunker",        -mc_va,             True,  False),
                    ("Total bunker",      -tb_va,             True,  True),
                    ("Port charges",      -pc_va,             True,  False),
                    ("Insurance + other", -(ins_va + oth_va), True,  False),
                    ("Total expenses",    -exp_va,            True,  True),
                ]
                rows_html_va = ""
                for lbl_va, val_va, is_cost_va, is_sub_va in pl_items_va:
                    vc_va = (
                        "#166534" if val_va > 0 and not is_cost_va
                        else "#991b1b" if val_va < 0
                        else "#1a3a5c"
                    )
                    fw_va  = "600" if is_sub_va else "400"
                    bdr_va = (
                        "border-top:0.5px solid #e2e8f0;padding-top:4px;"
                        if is_sub_va else ""
                    )
                    sg_va = "+" if val_va > 0 else ""
                    rows_html_va += (
                        f"<tr style='{bdr_va}'>"
                        f"<td style='padding:2px 0;font-size:11px;"
                        f"color:#475569;font-weight:{fw_va}'>{lbl_va}</td>"
                        f"<td style='text-align:right;font-size:11px;"
                        f"color:{vc_va};font-weight:{fw_va}'>"
                        f"{sg_va}${abs(val_va):,.0f}</td></tr>"
                    )
                pl_c_va = "#166534" if pl_va >= 0 else "#991b1b"
                vs_c_va = "#166534" if vs_avg_va >= 0 else "#991b1b"
                vs_s_va = "+" if vs_avg_va >= 0 else ""
                st.markdown(
                    f"<table style='width:100%;border-collapse:collapse'>"
                    f"{rows_html_va}"
                    f"<tr style='border-top:2px solid #1a3a5c'>"
                    f"<td style='padding:4px 0;font-size:13px;"
                    f"font-weight:700;color:#1a3a5c'>Net profit / loss</td>"
                    f"<td style='text-align:right;font-size:15px;"
                    f"font-weight:700;color:{pl_c_va}'>"
                    f"{'+'if pl_va>=0 else ''}${pl_va:,.0f}</td></tr>"
                    f"<tr><td style='font-size:10px;color:#94a3b8;padding:2px 0'>"
                    f"TCE this voyage</td>"
                    f"<td style='text-align:right;font-size:10px;"
                    f"color:#475569'>${ppd_va:,.0f}/day</td></tr>"
                    f"<tr><td style='font-size:10px;color:#94a3b8'>"
                    f"vs programme avg</td>"
                    f"<td style='text-align:right;font-size:10px;"
                    f"color:{vs_c_va}'>{vs_s_va}${vs_avg_va:,.0f}/day</td></tr>"
                    f"<tr><td style='font-size:10px;color:#94a3b8'>"
                    f"Total voyage days</td>"
                    f"<td style='text-align:right;font-size:10px;"
                    f"color:#475569'>{td_va:.1f} days</td></tr>"
                    f"</table>",
                    unsafe_allow_html=True
                )

            with wf_col_r:
                st.markdown(
                    "<div style='font-size:10px;font-weight:600;color:#64748b;"
                    "text-transform:uppercase;letter-spacing:.06em;"
                    "margin-bottom:8px'>Waterfall chart</div>",
                    unsafe_allow_html=True
                )
                wf_labels_va = [
                    'Gross', 'Brokerage', 'Net inc',
                    'Hire', 'Bunker', 'Port', 'Ins+oth', 'Profit'
                ]
                wf_values_va = [
                    g_va, -br_va, ni_va,
                    -ch_va, -tb_va, -pc_va,
                    -(ins_va+oth_va), pl_va
                ]
                wf_colors_va = [
                    '#378ADD', '#ef4444', '#378ADD',
                    '#ef4444', '#ef4444', '#ef4444', '#ef4444',
                    '#22c55e' if pl_va >= 0 else '#ef4444'
                ]
                fig_wf_va = go.Figure(go.Bar(
                    x=wf_labels_va,
                    y=[abs(v) for v in wf_values_va],
                    marker_color=wf_colors_va,
                    text=[f"${abs(v)/1000:.0f}k" for v in wf_values_va],
                    textposition='outside',
                    textfont=dict(size=9),
                    showlegend=False,
                    hovertemplate='%{x}<br>$%{y:,.0f}<extra></extra>',
                ))
                fig_wf_va.update_layout(
                    height=260, margin=dict(l=0, r=0, t=16, b=0),
                    plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(tickfont=dict(size=9)),
                    yaxis=dict(visible=False),
                )
                st.plotly_chart(fig_wf_va, use_container_width=True)

                st.markdown(
                    f"<div style='display:flex;gap:6px;flex-wrap:wrap;"
                    f"margin-top:4px'>"
                    f"<span style='font-size:10px;background:#e6f1fb;"
                    f"color:#0c447c;padding:2px 8px;border-radius:10px'>"
                    f"Load nav: ${load_meta['nav']:,}</span>"
                    f"<span style='font-size:10px;background:#e6f1fb;"
                    f"color:#0c447c;padding:2px 8px;border-radius:10px'>"
                    f"Disch nav: ${disch_meta['nav']:,}</span>"
                    f"<span style='font-size:10px;background:#fef3c7;"
                    f"color:#92400e;padding:2px 8px;border-radius:10px'>"
                    f"Cong: "
                    f"{load_meta['cong']+disch_meta['cong']:.1f}d</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

        # ════════════════════════════════════════════════════════════════
        # ZONE 2 — Full-width sea route map
        # ════════════════════════════════════════════════════════════════
        st.markdown("---")
        st.markdown(
            f"<div style='font-size:12px;font-weight:600;color:#1a3a5c;"
            f"margin-bottom:6px'>"
            f"Sea route — {new_load} → {new_disch}"
            f"<span style='font-size:11px;font-weight:400;color:#64748b;"
            f"margin-left:8px'>"
            f"Laden: {laden_nm:,.0f} NM · {laden_days:.1f} days  |  "
            f"Ballast: {ballast_nm:,.0f} NM · {ballast_days:.1f} days"
            f"</span></div>",
            unsafe_allow_html=True
        )

        map_traces_va = []

        # Ballast leg — grey dashed line from prev port to load
        if sel_idx > 0 and ballast_nm > 0:
            prev_dest_va = legs_list[sel_idx-1].get('dest_port', '')
            if (prev_dest_va in port_coords_va
                    and new_load in port_coords_va):
                bl, blo   = port_coords_va[prev_dest_va]
                ll2, llo2 = port_coords_va[new_load]
                b_lats, b_lons = _geodesic(bl, blo, ll2, llo2)
                map_traces_va.append(go.Scattermap(
                    lat=b_lats, lon=b_lons, mode='lines',
                    line=dict(width=2, color='#94a3b8'),
                    opacity=0.6, hoverinfo='skip', showlegend=False,
                ))
                bm = len(b_lats) // 2
                map_traces_va.append(go.Scattermap(
                    lat=[b_lats[bm]], lon=[b_lons[bm]],
                    mode='markers+text',
                    marker=dict(size=1, color='rgba(0,0,0,0)'),
                    text=[f"Ballast {ballast_nm:,.0f} NM"],
                    textfont=dict(size=10, color='#64748b'),
                    textposition='top center',
                    hoverinfo='skip', showlegend=False,
                ))

        # Laden leg — main route arc
        if new_load in port_coords_va and new_disch in port_coords_va:
            ll_va, llo_va  = port_coords_va[new_load]
            dl_va, dlo_va  = port_coords_va[new_disch]
            l_lats, l_lons = _geodesic(ll_va, llo_va, dl_va, dlo_va)
            route_col_va   = '#f59e0b' if pl_va >= 0 else '#ef4444'

            map_traces_va.append(go.Scattermap(
                lat=l_lats, lon=l_lons, mode='lines',
                line=dict(width=5, color=route_col_va),
                hovertext=(
                    f"<b>Voyage {sel_idx+1} — {new_comm}</b><br>"
                    f"{new_load} → {new_disch}<br>"
                    f"Distance: {laden_nm:,.0f} NM<br>"
                    f"Sailing: {laden_days:.1f} days<br>"
                    f"Cargo: {new_cargo:,.0f} MT @ ${new_rate}/MT<br>"
                    f"P&L: {'+'if pl_va>=0 else ''}${pl_va:,.0f}"
                ),
                hoverinfo='text', showlegend=False,
            ))

            # Distance label at arc midpoint
            lm_va_idx = len(l_lats) // 2
            map_traces_va.append(go.Scattermap(
                lat=[l_lats[lm_va_idx]], lon=[l_lons[lm_va_idx]],
                mode='markers+text',
                marker=dict(size=1, color='rgba(0,0,0,0)'),
                text=[f"{laden_nm:,.0f} NM · {laden_days:.1f} days"],
                textfont=dict(size=11, color='#92400e'),
                textposition='top center',
                hoverinfo='skip', showlegend=False,
            ))

            # Vessel marker at destination
            map_traces_va.append(go.Scattermap(
                lat=[l_lats[-1]], lon=[l_lons[-1]],
                mode='markers',
                marker=dict(size=18, color='#f59e0b'),
                hovertext='Vessel position',
                hoverinfo='text', showlegend=False,
            ))

        # Programme route ports as grey background dots
        route_ports_va = set()
        for lg in legs_list:
            route_ports_va.add(lg.get('origin_port', ''))
            route_ports_va.add(lg.get('dest_port', ''))
        route_ports_va.discard('')
        valid_rp_va = [p for p in route_ports_va if p in port_coords_va]
        if valid_rp_va:
            map_traces_va.append(go.Scattermap(
                lat=[port_coords_va[p][0] for p in valid_rp_va],
                lon=[port_coords_va[p][1] for p in valid_rp_va],
                mode='markers',
                marker=dict(size=5, color='#94a3b8', opacity=0.4),
                hoverinfo='skip', showlegend=False,
            ))

        # Load and discharge port markers
        for pname_va, pcolor_va, psize_va in [
            (new_load,  '#1d4ed8', 16),
            (new_disch, '#dc2626', 16),
        ]:
            if pname_va in port_coords_va:
                meta_va          = _get_meta(pname_va)
                plat_va, plon_va = port_coords_va[pname_va]
                map_traces_va.append(go.Scattermap(
                    lat=[plat_va], lon=[plon_va],
                    mode='markers+text',
                    marker=dict(size=psize_va, color=pcolor_va),
                    text=[pname_va],
                    textposition='top right',
                    textfont=dict(size=11, color='#1e293b'),
                    hovertemplate=(
                        f"<b>{pname_va}</b><br>"
                        f"Country: {meta_va['country']}<br>"
                        f"Port nav charge: ${meta_va['nav']:,}<br>"
                        f"Avg congestion: {meta_va['cong']:.1f} days"
                        f"<extra></extra>"
                    ),
                    showlegend=False,
                ))

        # Map centre between load and discharge
        if new_load in port_coords_va and new_disch in port_coords_va:
            clat_va = (port_coords_va[new_load][0]
                       + port_coords_va[new_disch][0]) / 2
            clon_va = (port_coords_va[new_load][1]
                       + port_coords_va[new_disch][1]) / 2
        else:
            clat_va, clon_va = 10.0, 105.0

        fig_map_va = go.Figure(data=map_traces_va)
        fig_map_va.update_layout(
            mapbox=dict(
                style='carto-positron', zoom=3.5,
                center=dict(lat=clat_va, lon=clon_va),
            ),
            height=520, margin=dict(l=0, r=0, t=0, b=0),
            paper_bgcolor='rgba(0,0,0,0)', showlegend=False,
            hoverlabel=dict(bgcolor='white', bordercolor='#1e293b',
                            font=dict(size=12)),
        )
        st.plotly_chart(fig_map_va, use_container_width=True)

        # Map legend strip
        ml1, ml2, ml3, ml4, ml5 = st.columns(5)
        ml1.markdown(
            "<div style='display:flex;align-items:center;gap:4px'>"
            "<div style='width:10px;height:10px;border-radius:50%;"
            "background:#1d4ed8'></div>"
            "<span style='font-size:11px;color:#475569'>Load port</span>"
            "</div>", unsafe_allow_html=True
        )
        ml2.markdown(
            "<div style='display:flex;align-items:center;gap:4px'>"
            "<div style='width:10px;height:10px;border-radius:50%;"
            "background:#dc2626'></div>"
            "<span style='font-size:11px;color:#475569'>Discharge port"
            "</span></div>", unsafe_allow_html=True
        )
        ml3.markdown(
            "<div style='display:flex;align-items:center;gap:4px'>"
            "<div style='width:22px;height:3px;background:#f59e0b'></div>"
            "<span style='font-size:11px;color:#475569'>Laden route</span>"
            "</div>", unsafe_allow_html=True
        )
        ml4.markdown(
            "<div style='display:flex;align-items:center;gap:4px'>"
            "<div style='width:22px;height:2px;background:#94a3b8'></div>"
            "<span style='font-size:11px;color:#475569'>Ballast leg</span>"
            "</div>", unsafe_allow_html=True
        )
        ml5.markdown(
            f"<div style='text-align:right'>"
            f"<a href='https://www.marinetraffic.com/en/ais/home/"
            f"centerx:{clon_va:.1f}/centery:{clat_va:.1f}/zoom:6' "
            f"style='font-size:11px;color:#0369a1;text-decoration:none'>"
            f"View on MarineTraffic ↗</a></div>",
            unsafe_allow_html=True
        )

# ─── TAB 7: VOYAGE JOURNEY (animated vessel) ─────────────────────────────────
with tabs[5]:
    if 'analysis' not in st.session_state:
        st.info(
            "Run the simulation first — click 🚀 Run Simulation "
            "in the sidebar to generate voyage programmes."
        )
    else:
        import math as _math

        analysis_vj  = st.session_state['analysis']
        vessel_vj    = st.session_state.get('vessel')
        top_progs_vj = analysis_vj['top_programmes']

        intra_vj, ports_vj, dist_matrix_vj, legs_vj = load_data(DATA_PATH)
        port_coords_vj  = dict(zip(
            ports_vj['Port'], zip(ports_vj['Lat'], ports_vj['Lon'])
        ))
        port_country_vj = dict(zip(ports_vj['Port'], ports_vj['Country']))

        from data.port_charges import PORT_CHARGES, PORT_CHARGES_DEFAULT

        def _get_meta_vj(pname):
            pc = PORT_CHARGES.get(pname, PORT_CHARGES_DEFAULT)
            return {
                'nav':     pc.get('nav', 10000),
                'cong':    pc.get('cong_mean', 1.0),
                'country': port_country_vj.get(pname, '—'),
            }

        def _calc_nm_vj(pa, pb):
            if pa not in port_coords_vj or pb not in port_coords_vj:
                return 0
            la1, lo1 = port_coords_vj[pa]
            la2, lo2 = port_coords_vj[pb]
            R = 3440.065
            dlat = _math.radians(la2 - la1)
            dlon = _math.radians(lo2 - lo1)
            a = (_math.sin(dlat/2)**2
                 + _math.cos(_math.radians(la1))
                 * _math.cos(_math.radians(la2))
                 * _math.sin(dlon/2)**2)
            straight = 2 * R * _math.asin(_math.sqrt(a))
            c1 = port_country_vj.get(pa, '')
            c2 = port_country_vj.get(pb, '')
            f = (1.45 if c1 == 'Indonesia' and c2 == 'Indonesia'
                 else 1.40 if 'Bangladesh' in [c1, c2]
                 else 1.35 if 'Philippines' in [c1, c2]
                 else 1.25)
            return round(straight * f)

        def _geodesic_vj(lat1, lon1, lat2, lon2, n=60):
            if lat1 == lat2 and lon1 == lon2:
                return [lat1], [lon1]
            la1 = _math.radians(lat1); lo1 = _math.radians(lon1)
            la2 = _math.radians(lat2); lo2 = _math.radians(lon2)
            d = 2 * _math.asin(_math.sqrt(
                _math.sin((la2-la1)/2)**2
                + _math.cos(la1) * _math.cos(la2)
                * _math.sin((lo2-lo1)/2)**2
            ))
            if d == 0:
                return [lat1, lat2], [lon1, lon2]
            lats, lons = [], []
            for i in range(n + 1):
                f  = i / n
                A  = _math.sin((1-f)*d) / _math.sin(d)
                B  = _math.sin(f*d)     / _math.sin(d)
                x  = (A*_math.cos(la1)*_math.cos(lo1)
                      + B*_math.cos(la2)*_math.cos(lo2))
                y  = (A*_math.cos(la1)*_math.sin(lo1)
                      + B*_math.cos(la2)*_math.sin(lo2))
                z  = A*_math.sin(la1) + B*_math.sin(la2)
                lats.append(_math.degrees(_math.atan2(
                    z, _math.sqrt(x*x + y*y))))
                lons.append(_math.degrees(_math.atan2(y, x)))
            return lats, lons

        # ── Programme selector ────────────────────────────────────────
        hc1, hc2, hc3 = st.columns([3, 1, 1])
        with hc1:
            prog_opts_vj = [
                f"#{p['rank']}  —  ${p['total_profit']:,.0f}  —  "
                f"{p['n_voyages']} voyages  —  TCE ${p['avg_tce']:,.0f}/day"
                for p in top_progs_vj
            ]
            sel_prog_vj = st.selectbox(
                "Programme",
                range(len(top_progs_vj)),
                format_func=lambda i: prog_opts_vj[i],
                key='vj_prog_sel',
                label_visibility='collapsed',
            )
        prog_vj      = top_progs_vj[sel_prog_vj]
        legs_list_vj = prog_vj['legs']

        with hc2:
            st.metric("Annual Profit", f"${prog_vj['total_profit']:,.0f}")
        with hc3:
            try:
                eb_vj = export_programme_to_excel(
                    prog_vj, vessel_vj, prog_vj['rank']
                )
                st.download_button(
                    label="📥 Export Excel", data=eb_vj,
                    file_name=f"voyage_journey_{prog_vj['rank']}.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                    use_container_width=True,
                )
            except Exception:
                pass


        # -- Programme summary strip --
        ps1, ps2, ps3, ps4 = st.columns(4)
        ps1.metric("Voyages",     prog_vj['n_voyages'])
        ps2.metric("TCE",         f"${prog_vj['avg_tce']:,.0f}/day")
        ps3.metric("Utilisation", f"{prog_vj['utilisation_pct']:.1f}%")
        ps4.metric("Algorithm",   prog_vj.get('algorithm', 'mc').upper())

        st.markdown("---")

        # ═══════════════════════════════════════════════════════════════
        # ZONE 2 — Pure HTML/JS animated vessel journey (zero Python)
        # ═══════════════════════════════════════════════════════════════

        st.markdown(
            "<div style='margin-bottom:12px;padding:12px 18px;"
            "background:linear-gradient(135deg,#1a3a5c,#2d5986);"
            "border-radius:10px;display:flex;align-items:center;"
            "justify-content:space-between;flex-wrap:wrap;gap:8px'>"
            "<div>"
            "<span style='font-size:14px;font-weight:700;color:white'>"
            "🚢 Annual Vessel Journey</span>"
            "<span style='font-size:11px;color:rgba(255,255,255,0.7);"
            "margin-left:10px'>"
            "Runs entirely in browser — zero lag</span>"
            "</div>"
            f"<span style='font-size:11px;color:rgba(255,255,255,0.85)'>"
            f"{prog_vj['n_voyages']} voyages · "
            f"${prog_vj['total_profit']:,.0f} profit · "
            f"TCE ${prog_vj['avg_tce']:,.0f}/day"
            f"</span></div>",
            unsafe_allow_html=True
        )

        # ── Pre-compute all data in Python → serialise to JSON ────────
        import math as _math2
        import json as _json2

        def _geo2(lat1, lon1, lat2, lon2, n=40):
            if lat1 == lat2 and lon1 == lon2:
                return [lat1, lat2], [lon1, lon2]
            r  = _math2.radians
            d2 = _math2.degrees
            la1,lo1,la2,lo2 = r(lat1),r(lon1),r(lat2),r(lon2)
            d = 2*_math2.asin(_math2.sqrt(
                _math2.sin((la2-la1)/2)**2
                + _math2.cos(la1)*_math2.cos(la2)
                  *_math2.sin((lo2-lo1)/2)**2
            ))
            if d == 0:
                return [lat1,lat2],[lon1,lon2]
            lats,lons=[],[]
            for i in range(n+1):
                f=i/n
                A=_math2.sin((1-f)*d)/_math2.sin(d)
                B=_math2.sin(f*d)/_math2.sin(d)
                x=A*_math2.cos(la1)*_math2.cos(lo1)+B*_math2.cos(la2)*_math2.cos(lo2)
                y=A*_math2.cos(la1)*_math2.sin(lo1)+B*_math2.cos(la2)*_math2.sin(lo2)
                z=A*_math2.sin(la1)+B*_math2.sin(la2)
                lats.append(d2(_math2.atan2(z,_math2.sqrt(x*x+y*y))))
                lons.append(d2(_math2.atan2(y,x)))
            return lats,lons

        def _nm2(pa, pb):
            if pa not in port_coords_vj or pb not in port_coords_vj:
                return 0
            la1,lo1=port_coords_vj[pa]
            la2,lo2=port_coords_vj[pb]
            R=3440.065
            r=_math2.radians
            dlat=r(la2-la1);dlon=r(lo2-lo1)
            a=(_math2.sin(dlat/2)**2
               +_math2.cos(r(la1))*_math2.cos(r(la2))
               *_math2.sin(dlon/2)**2)
            st2=2*R*_math2.asin(_math2.sqrt(a))
            c1=port_country_vj.get(pa,'')
            c2=port_country_vj.get(pb,'')
            f=(1.45 if c1=='Indonesia' and c2=='Indonesia'
               else 1.40 if 'Bangladesh' in [c1,c2]
               else 1.35 if 'Philippines' in [c1,c2]
               else 1.25)
            return round(st2*f)

        voyage_data_vj = []
        for leg in legs_list_vj:
            op = leg.get('origin_port','')
            dp = leg.get('dest_port','')
            pl = leg.get('profit_loss', leg.get('profit',0))
            if op in port_coords_vj and dp in port_coords_vj:
                lats2,lons2 = _geo2(*port_coords_vj[op], *port_coords_vj[dp])
            else:
                lats2,lons2 = [],[]
            voyage_data_vj.append({
                'from':  op,
                'to':    dp,
                'comm':  leg.get('commodity',''),
                'pl':    round(pl),
                'days':  round(leg.get('total_days',0),1),
                'nm':    _nm2(op,dp),
                'rate':  round(leg.get('freight_rate',0),2),
                'cargo': round(leg.get('cargo_mt',0)),
                'cum':   round(leg.get('cum_profit',0)),
                'lats':  lats2,
                'lons':  lons2,
                'olat':  port_coords_vj[op][0] if op in port_coords_vj else 0,
                'olon':  port_coords_vj[op][1] if op in port_coords_vj else 0,
                'dlat':  port_coords_vj[dp][0] if dp in port_coords_vj else 0,
                'dlon':  port_coords_vj[dp][1] if dp in port_coords_vj else 0,
            })

        all_ports_vj2 = [
            {'name': p,
             'lat':  port_coords_vj[p][0],
             'lon':  port_coords_vj[p][1]}
            for p in set(l.get('origin_port','') for l in legs_list_vj)
                     | set(l.get('dest_port','') for l in legs_list_vj)
            if p and p in port_coords_vj
        ]

        if all_ports_vj2:
            _clat2 = sum(p['lat'] for p in all_ports_vj2)/len(all_ports_vj2)
            _clon2 = sum(p['lon'] for p in all_ports_vj2)/len(all_ports_vj2)
        else:
            _clat2, _clon2 = 8.0, 108.0

        voyage_json2  = _json2.dumps(voyage_data_vj)
        ports_json2   = _json2.dumps(all_ports_vj2)
        centre_json2  = _json2.dumps({'lat': _clat2, 'lon': _clon2})

        # ── Pure HTML/JS component ─────────────────────────────────────
        html_vj = f"""
<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
* {{ box-sizing:border-box; margin:0; padding:0; font-family:sans-serif }}
body {{ background:transparent }}
#map {{ width:100%; height:480px; border-radius:10px;
        border:1px solid #e2e8f0; overflow:hidden }}
.controls {{ display:flex; align-items:center; gap:8px;
             margin:10px 0 8px; flex-wrap:wrap }}
.btn-primary {{ background:#1a3a5c; color:white; border:none;
                padding:8px 20px; border-radius:8px; font-size:13px;
                font-weight:600; cursor:pointer; min-width:90px }}
.btn-primary:hover {{ background:#2d5986 }}
.btn-sec {{ background:white; color:#1a3a5c;
            border:1px solid #e2e8f0; padding:8px 14px;
            border-radius:8px; font-size:13px; cursor:pointer }}
.btn-sec:hover {{ background:#f8fafc }}
select {{ padding:7px 12px; border-radius:8px;
          border:1px solid #e2e8f0; font-size:13px;
          background:white; color:#1a3a5c; cursor:pointer }}
input[type=range] {{ flex:1; min-width:100px; accent-color:#1a3a5c }}
.badge {{ background:linear-gradient(135deg,#1a3a5c,#2d5986);
          color:white; padding:6px 14px; border-radius:8px;
          font-size:12px; font-weight:600; white-space:nowrap }}
.stats {{ display:grid; grid-template-columns:repeat(6,1fr);
          gap:8px; margin-top:8px }}
.stat {{ background:#f8fafc; border-radius:8px; padding:8px 10px;
         border:1px solid #e2e8f0 }}
.stat-lbl {{ font-size:10px; color:#94a3b8; text-transform:uppercase;
             letter-spacing:.04em }}
.stat-val {{ font-size:13px; font-weight:600; color:#1a3a5c;
             margin-top:2px; white-space:nowrap; overflow:hidden;
             text-overflow:ellipsis }}
.leg-row {{ display:flex; align-items:center; gap:6px; padding:5px 8px;
            border-radius:6px; cursor:pointer; font-size:11px;
            transition:background .1s; border:1px solid transparent }}
.leg-row:hover {{ background:#f1f5f9 }}
.leg-row.active {{ background:#eaf3de; border-color:#639922 }}
.tag {{ font-size:10px; padding:2px 7px; border-radius:10px;
        font-weight:500; white-space:nowrap }}
.tag-g {{ background:#dcfce7; color:#166534 }}
.tag-r {{ background:#fee2e2; color:#991b1b }}
.tag-b {{ background:#e6f1fb; color:#0c447c }}
.legend {{ display:flex; gap:14px; flex-wrap:wrap;
           margin-top:8px; align-items:center; font-size:11px;
           color:#64748b }}
.leg-dot {{ width:12px; height:12px; border-radius:50%;
            display:inline-block; margin-right:4px }}
.leg-line {{ width:22px; height:3px; display:inline-block;
             margin-right:4px; vertical-align:middle }}
</style></head>
<body>
<div id="map"></div>
<div class="controls">
  <button class="btn-primary" id="btnPlay">▶ Play</button>
  <button class="btn-sec"     id="btnReset">↺ Reset</button>
  <select id="spdSel">
    <option value="120">Slow</option>
    <option value="60" selected>Normal</option>
    <option value="30">Fast</option>
    <option value="15">Very Fast</option>
  </select>
  <input type="range" id="scrub" min="0" max="100" value="0" step="1">
  <span class="badge" id="voyBadge">V 1 / —</span>
</div>
<div class="stats">
  <div class="stat"><div class="stat-lbl">Route</div>
    <div class="stat-val" id="sRoute">—</div></div>
  <div class="stat"><div class="stat-lbl">Commodity</div>
    <div class="stat-val" id="sComm">—</div></div>
  <div class="stat"><div class="stat-lbl">Distance</div>
    <div class="stat-val" id="sNm">—</div></div>
  <div class="stat"><div class="stat-lbl">Voyage P&L</div>
    <div class="stat-val" id="sPl">—</div></div>
  <div class="stat"><div class="stat-lbl">Cumulative</div>
    <div class="stat-val" id="sCum">$0</div></div>
  <div class="stat"><div class="stat-lbl">Days elapsed</div>
    <div class="stat-val" id="sDays">0 d</div></div>
</div>
<div style="margin-top:10px;font-size:11px;color:#64748b;margin-bottom:4px">
  Voyage schedule — click to jump
</div>
<div id="legList"
     style="max-height:180px;overflow-y:auto;display:flex;
            flex-direction:column;gap:2px;padding-right:4px"></div>
<div class="legend">
  <span><span class="leg-dot" style="background:#f59e0b"></span>Vessel</span>
  <span><span class="leg-dot" style="background:#1d4ed8"></span>Load port</span>
  <span><span class="leg-dot" style="background:#dc2626"></span>Discharge</span>
  <span><span class="leg-line" style="background:#22c55e"></span>Profitable</span>
  <span><span class="leg-line" style="background:#ef4444"></span>Loss leg</span>
  <a href="https://www.marinetraffic.com" target="_blank"
     style="margin-left:auto;color:#0369a1;text-decoration:none">
    MarineTraffic ↗</a>
</div>
<script src="https://cdn.jsdelivr.net/npm/plotly.js-dist@2.26.0/plotly.min.js"></script>
<script>
const VOYAGES = {voyage_json2};
const ALL_PORTS = {ports_json2};
const CENTRE = {centre_json2};
const globalFrames = [];
VOYAGES.forEach((v, vi) => {{
  if (!v.lats.length) return;
  const step = Math.max(1, Math.floor(v.lats.length / 30));
  for (let si = 0; si < v.lats.length; si += step) {{
    globalFrames.push({{ vi, si, lat: v.lats[si], lon: v.lons[si] }});
  }}
  const last = v.lats.length - 1;
  if (globalFrames[globalFrames.length-1].si !== last)
    globalFrames.push({{ vi, si: last, lat: v.lats[last], lon: v.lons[last] }});
}});
const traces = [];
traces.push({{
  type:'scattermapbox',
  lat: ALL_PORTS.map(p => p.lat),
  lon: ALL_PORTS.map(p => p.lon),
  mode:'markers',
  marker:{{size:6, color:'#64748b', opacity:0.4}},
  hoverinfo:'skip', showlegend:false, name:'bg_ports'
}});
VOYAGES.forEach((v, vi) => {{
  if (!v.lats.length) return;
  const col = v.pl >= 0 ? '#22c55e' : '#ef4444';
  traces.push({{
    type:'scattermapbox',
    lat: v.lats, lon: v.lons, mode:'lines',
    line:{{width:2, color:col}}, opacity:0.18,
    hoverinfo:'skip', showlegend:false, name:`arc_${{vi}}`
  }});
}});
ALL_PORTS.forEach(p => {{
  traces.push({{
    type:'scattermapbox',
    lat:[p.lat], lon:[p.lon],
    mode:'markers+text',
    marker:{{size:7, color:'#475569', opacity:0.6}},
    text:[p.name], textposition:'top right',
    textfont:{{size:9, color:'#334155'}},
    hovertemplate:`<b>${{p.name}}</b><extra></extra>`,
    showlegend:false, name:`lbl_${{p.name}}`
  }});
}});
const ACTIVE_ARC_IDX = traces.length;
traces.push({{
  type:'scattermapbox',
  lat:[], lon:[], mode:'lines',
  line:{{width:6, color:'#f59e0b'}},
  opacity:1, hoverinfo:'skip', showlegend:false, name:'active_arc'
}});
const VESSEL_IDX = traces.length;
const f0 = globalFrames[0] || {{vi:0,si:0,lat:CENTRE.lat,lon:CENTRE.lon}};
traces.push({{
  type:'scattermapbox',
  lat:[f0.lat], lon:[f0.lon], mode:'markers',
  marker:{{size:20, color:'#f59e0b'}},
  hovertext:'🚢 Vessel', hoverinfo:'text',
  showlegend:false, name:'vessel'
}});
const LOAD_IDX = traces.length;
traces.push({{
  type:'scattermapbox',
  lat:[VOYAGES[0]?.olat||0], lon:[VOYAGES[0]?.olon||0],
  mode:'markers+text',
  marker:{{size:16, color:'#1d4ed8'}},
  text:[VOYAGES[0]?.from||''], textposition:'top right',
  textfont:{{size:11, color:'#1e293b'}},
  hoverinfo:'skip', showlegend:false, name:'load_port'
}});
const DISCH_IDX = traces.length;
traces.push({{
  type:'scattermapbox',
  lat:[VOYAGES[0]?.dlat||0], lon:[VOYAGES[0]?.dlon||0],
  mode:'markers+text',
  marker:{{size:16, color:'#dc2626'}},
  text:[VOYAGES[0]?.to||''], textposition:'top right',
  textfont:{{size:11, color:'#1e293b'}},
  hoverinfo:'skip', showlegend:false, name:'disch_port'
}});
const layout = {{
  mapbox:{{
    style:'carto-positron',
    zoom:4.2,
    center:{{lat:CENTRE.lat, lon:CENTRE.lon}}
  }},
  height:480,
  margin:{{l:0,r:0,t:0,b:0}},
  paper_bgcolor:'rgba(0,0,0,0)',
  showlegend:false,
  hoverlabel:{{bgcolor:'white',bordercolor:'#1e293b',font:{{size:12}}}}
}};
Plotly.newPlot('map', traces, layout, {{
  responsive:true,
  displayModeBar:false,
  scrollZoom:true
}});
const legList = document.getElementById('legList');
VOYAGES.forEach((v, vi) => {{
  const row = document.createElement('div');
  row.className = 'leg-row';
  row.id = `leg_${{vi}}`;
  const plTag = v.pl >= 0 ? 'tag-g' : 'tag-r';
  const sign  = v.pl >= 0 ? '+' : '';
  const fromS = v.from.split(' ')[0];
  const toS   = v.to.split(' ')[0];
  row.innerHTML = `
    <span style="min-width:20px;color:#94a3b8;font-size:10px">${{vi+1}}</span>
    <span style="flex:1;color:#1e293b;font-weight:500">
      ${{fromS}} → ${{toS}}</span>
    <span class="tag tag-b">${{v.comm}}</span>
    <span class="tag ${{plTag}}">${{sign}}$${{Math.abs(v.pl/1000).toFixed(0)}}k</span>
    <span style="color:#94a3b8;min-width:30px;text-align:right">
      ${{v.days}}d</span>`;
  row.addEventListener('click', () => jumpToVoyage(vi));
  legList.appendChild(row);
}});
let curFrame = 0, playing = false, timer = null;
const scrub = document.getElementById('scrub');
scrub.max = globalFrames.length - 1;
function updateStats(gf) {{
  const v = VOYAGES[gf.vi];
  if (!v) return;
  const frac = gf.si / Math.max(v.lats.length - 1, 1);
  const cumPl   = VOYAGES.slice(0, gf.vi).reduce((s,x)=>s+x.pl, 0)
                  + v.pl * frac;
  const cumDays = VOYAGES.slice(0, gf.vi).reduce((s,x)=>s+x.days, 0)
                  + v.days * frac;
  const fromS = v.from.split(' ')[0];
  const toS   = v.to.split(' ')[0];
  document.getElementById('sRoute').textContent = `${{fromS}} → ${{toS}}`;
  document.getElementById('sComm').textContent  = v.comm;
  document.getElementById('sNm').textContent    = `${{v.nm.toLocaleString()}} NM`;
  const plEl = document.getElementById('sPl');
  plEl.textContent  = (v.pl>=0?'+':'') + '$' + Math.abs(v.pl).toLocaleString();
  plEl.style.color  = v.pl >= 0 ? '#166534' : '#991b1b';
  const cumEl = document.getElementById('sCum');
  const cumR  = Math.round(cumPl);
  cumEl.textContent = (cumR>=0?'+':'') + '$' + Math.abs(cumR).toLocaleString();
  cumEl.style.color = cumPl >= 0 ? '#166534' : '#991b1b';
  document.getElementById('sDays').textContent =
    Math.round(cumDays) + ' d';
  document.getElementById('voyBadge').textContent =
    `V ${{gf.vi+1}} / ${{VOYAGES.length}}`;
  document.querySelectorAll('.leg-row').forEach((el, i) => {{
    el.classList.toggle('active', i === gf.vi);
  }});
  const activeRow = document.getElementById(`leg_${{gf.vi}}`);
  if (activeRow) activeRow.scrollIntoView({{block:'nearest'}});
}}
function renderFrame(fi) {{
  if (fi < 0 || fi >= globalFrames.length) return;
  const gf  = globalFrames[fi];
  const v   = VOYAGES[gf.vi];
  if (!v) return;
  const partLats = v.lats.slice(0, gf.si + 1);
  const partLons = v.lons.slice(0, gf.si + 1);
  Plotly.restyle('map', {{
    lat: [partLats],
    lon: [partLons]
  }}, [ACTIVE_ARC_IDX]);
  Plotly.restyle('map', {{
    lat: [[gf.lat]],
    lon: [[gf.lon]]
  }}, [VESSEL_IDX]);
  Plotly.restyle('map', {{
    lat: [[v.olat]], lon: [[v.olon]], text: [[v.from]]
  }}, [LOAD_IDX]);
  Plotly.restyle('map', {{
    lat: [[v.dlat]], lon: [[v.dlon]], text: [[v.to]]
  }}, [DISCH_IDX]);
  scrub.value = fi;
  updateStats(gf);
}}
function jumpToVoyage(vi) {{
  const fi = globalFrames.findIndex(f => f.vi === vi);
  if (fi >= 0) {{ curFrame = fi; renderFrame(curFrame); }}
}}
function advance() {{
  if (curFrame >= globalFrames.length - 1) {{
    playing = false; clearInterval(timer);
    document.getElementById('btnPlay').textContent = '▶ Play';
    return;
  }}
  curFrame++; renderFrame(curFrame);
}}
document.getElementById('btnPlay').addEventListener('click', function() {{
  if (playing) {{
    playing = false; clearInterval(timer);
    this.textContent = '▶ Play';
  }} else {{
    if (curFrame >= globalFrames.length - 1) curFrame = 0;
    playing = true;
    this.textContent = '⏸ Pause';
    const spd = parseInt(document.getElementById('spdSel').value);
    timer = setInterval(advance, spd);
  }}
}});
document.getElementById('btnReset').addEventListener('click', () => {{
  playing = false; clearInterval(timer);
  document.getElementById('btnPlay').textContent = '▶ Play';
  curFrame = 0; renderFrame(0);
}});
document.getElementById('scrub').addEventListener('input', e => {{
  if (playing) {{
    playing = false; clearInterval(timer);
    document.getElementById('btnPlay').textContent = '▶ Play';
  }}
  curFrame = parseInt(e.target.value);
  renderFrame(curFrame);
}});
document.getElementById('spdSel').addEventListener('change', function() {{
  if (playing) {{ clearInterval(timer); timer = setInterval(advance, parseInt(this.value)); }}
}});
renderFrame(0);
</script></body></html>
"""
        components.html(html_vj, height=920, scrolling=False)

